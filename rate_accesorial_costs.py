"""
Accessorial Costs Analysis Script

This script analyzes the rate.xlsx file from the "Accessorial costs" tab:
1. Removes first 2 rows
2. Extracts cost blocks with structure:
   - Cost name row
   - Multiplier row
   - Rate by row
   - Data row: Lane #, Currency, p/unit (or MIN -> Flat, p/unit), Apply if
   - Costs data rows
3. Creates AccessorialCost dataclass objects to store all conditions
"""

import pandas as pd
import openpyxl
import os
import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict


@dataclass
class LaneData:
    """Represents data for a single lane within an accessorial cost.
    
    Note on flat pricing:
    - price_flat_min: Contains the flat price. For plain flat prices (no MIN indicator),
                      this holds the value but has_min_flat will be False in the parent AccessorialCost.
                      For MIN Flat prices, has_min_flat will be True.
    - price_flat_max: Contains the MAX flat price when MAX indicator exists.
    
    Note on percentage pricing:
    - percentage: Contains the percentage value (e.g., 12 for 12%) when is_percentage=True
    - valid_from/valid_to: Date range for when this lane's pricing is valid
    """
    lane_number: int
    currency: Optional[str] = None
    price_flat_min: Optional[float] = None       # Flat price (plain) or MIN Flat price (if has_min_flat)
    price_flat_max: Optional[float] = None       # MAX Flat price (if has_max_flat)
    price_per_unit: Optional[float] = None
    percentage: Optional[float] = None           # Percentage value for % over costs
    applies_if: Optional[str] = None
    valid_from: Optional[str] = None             # Valid From date
    valid_to: Optional[str] = None               # Valid To date
    
    def to_dict(self) -> dict:
        return {
            "Lane #": self.lane_number,
            "Currency": self.currency or "",
            "Price Flat MIN": self.price_flat_min if self.price_flat_min is not None else "",
            "Price Flat MAX": self.price_flat_max if self.price_flat_max is not None else "",
            "Price per unit": self.price_per_unit if self.price_per_unit is not None else "",
            "Percentage": self.percentage if self.percentage is not None else "",
            "Applies If": self.applies_if or "",
            "Valid From": self.valid_from or "",
            "Valid To": self.valid_to or ""
        }


@dataclass
class AccessorialCost:
    """Represents an accessorial cost type with its conditions and pricing info."""
    name: str                                    # e.g., "Dangerous Goods", "Fuel Surcharge"
    rate_by: Optional[str] = None                # e.g., "Quantity/Container", "per shipment"
    multiplier: Optional[str] = None             # e.g., "per shipment", "per kg" (optional)
    has_min_flat: bool = False                   # True if has MIN + Flat pricing
    has_max_flat: bool = False                   # True if has MAX + Flat pricing
    is_percentage: bool = False                  # True if cost is % over other costs
    applied_over_costs: List[str] = field(default_factory=list)  # List of cost names this % applies over
    lanes: List['LaneData'] = field(default_factory=list)  # All lane data
    start_row: Optional[int] = None              # Row where this cost block starts
    
    def has_conditions(self) -> bool:
        """Check if any lane has an applies_if condition."""
        return any(lane.applies_if for lane in self.lanes)
    
    def has_date_validity(self) -> bool:
        """Check if any lane has date validity (Valid From/To)."""
        return any(lane.valid_from or lane.valid_to for lane in self.lanes)
    
    def get_price_type_description(self) -> str:
        """Get description of pricing type."""
        if self.is_percentage:
            return "% over costs"
        parts = []
        if self.has_min_flat:
            parts.append("MIN Flat")
        if self.has_max_flat:
            parts.append("MAX Flat")
        parts.append("per unit")
        return " + ".join(parts)
    
    def to_dict(self) -> dict:
        """Convert to dictionary for DataFrame export (one row per lane)."""
        return {
            "Cost Name": self.name,
            "Rate By": self.rate_by or "",
            "Multiplier": self.multiplier or "",
            "Price Type": self.get_price_type_description(),
            "Is Percentage": "Yes" if self.is_percentage else "No",
            "Applied Over": ", ".join(self.applied_over_costs) if self.applied_over_costs else "",
            "Has MIN Flat": "Yes" if self.has_min_flat else "No",
            "Has MAX Flat": "Yes" if self.has_max_flat else "No",
            "Num Lanes": len(self.lanes),
            "Has Conditions": "Yes" if self.has_conditions() else "No",
            "Has Date Validity": "Yes" if self.has_date_validity() else "No"
        }
    
    def to_lane_dicts(self) -> List[dict]:
        """Convert to list of dictionaries (one per lane) for detailed export."""
        result = []
        for lane in self.lanes:
            row = {
                "Cost Name": self.name,
                "Rate By": self.rate_by or "",
                "Multiplier": self.multiplier or "",
                "Price Type": self.get_price_type_description(),
                "Is Percentage": "Yes" if self.is_percentage else "No",
                "Applied Over": ", ".join(self.applied_over_costs) if self.applied_over_costs else "",
                "Has MIN Flat": "Yes" if self.has_min_flat else "No",
                "Has MAX Flat": "Yes" if self.has_max_flat else "No",
                "Lane #": lane.lane_number,
                "Currency": lane.currency or "",
                "Price Flat MIN": lane.price_flat_min if lane.price_flat_min is not None else "",
                "Price Flat MAX": lane.price_flat_max if lane.price_flat_max is not None else "",
                "Price per unit": lane.price_per_unit if lane.price_per_unit is not None else "",
                "Percentage": lane.percentage if lane.percentage is not None else "",
                "Applies If": lane.applies_if or "",
                "Valid From": lane.valid_from or "",
                "Valid To": lane.valid_to or ""
            }
            result.append(row)
        return result


def get_agreement_number(file_path):
    """
    Extract Agreement number from "General info" tab of a rate card file.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        str: Agreement number or None if not found
    """
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    agreement_number = None
    try:
        workbook_info = openpyxl.load_workbook(full_path, data_only=True)
        if "General info" in workbook_info.sheetnames:
            general_info_sheet = workbook_info["General info"]
            # Find row with "Agreement number" in column A
            for row in general_info_sheet.iter_rows(min_col=1, max_col=2):
                cell_a = row[0]
                cell_b = row[1] if len(row) > 1 else None
                if cell_a.value and "Agreement number" in str(cell_a.value):
                    if cell_b and cell_b.value:
                        agreement_number = str(cell_b.value).strip()
                    break
        workbook_info.close()
    except Exception as e:
        print(f"   Warning: Could not extract Agreement number: {e}")
    
    return agreement_number


def load_accessorial_costs_sheet(file_path):
    """Load the Accessorial costs sheet from the Excel file using openpyxl."""
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    
    # Try different possible sheet names
    possible_names = ["Accessorial costs", "Accessorial Costs", "accessorial costs", "Accessorial"]
    sheet = None
    sheet_name_found = None
    
    for name in possible_names:
        if name in workbook.sheetnames:
            sheet = workbook[name]
            sheet_name_found = name
            break
    
    if sheet is None:
        available_sheets = workbook.sheetnames
        raise ValueError(f"Accessorial costs sheet not found. Available sheets: {available_sheets}")
    
    print(f"Loaded '{sheet_name_found}' sheet from {file_path}")
    print(f"   Total rows: {sheet.max_row}")
    print(f"   Total columns: {sheet.max_column}")
    
    return sheet, workbook


def get_row_values(sheet, row_number):
    """Get all values from a specific row (1-indexed)."""
    row_values = []
    for cell in sheet[row_number]:
        row_values.append(cell.value)
    return row_values


def get_cell_value(sheet, row, col):
    """Get cell value at specific row and column (1-indexed)."""
    return sheet.cell(row=row, column=col).value


def clean_text(value):
    """Clean and strip text value."""
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def get_first_non_empty_value(row_values):
    """Get the first non-None, non-empty value from a row."""
    for val in row_values:
        if val is not None:
            text = str(val).strip()
            if text:
                return text
    return None


def get_first_non_empty_index(row_values):
    """Get the index of the first non-None, non-empty value from a row."""
    for i, val in enumerate(row_values):
        if val is not None:
            text = str(val).strip()
            if text:
                return i
    return None


def is_cost_name_row(row_values, row_idx):
    """
    Check if this row contains a cost name.
    Cost name rows typically have a non-empty value that is NOT a number
    and doesn't start with common data keywords.
    
    Note: The cost name may NOT be in column 0 - it could be in column B due to merged cells.
    
    Examples of valid cost names:
    - "Cancellation Fee (Prior to dispatch)"
    - "Dangerous Goods"
    - "Pickup Fee"
    """
    # Get first non-empty value (not necessarily column 0!)
    first_str = get_first_non_empty_value(row_values)
    if first_str is None:
        return False
    
    first_lower = first_str.lower()
    
    # Skip if it's a number (Lane #)
    try:
        float(first_str)
        return False
    except ValueError:
        pass
    
    # Skip common data row indicators (case insensitive)
    skip_keywords = ['lane', 'currency', 'flat', 'p/unit', 'min', 'apply if', 'applies if', 'rate by', 'rate by:', 'multiplier', 'per ']
    for keyword in skip_keywords:
        if first_lower.startswith(keyword):
            return False
    
    # Skip if it's just empty or whitespace
    if not first_str or len(first_str) < 2:
        return False
    
    # Accept cost names - they can start with "(" or be regular text
    # Cost names are typically descriptive text that doesn't match data patterns
    if first_str.startswith('(') or (len(first_str) > 2 and not first_str.isdigit()):
        return True
    
    return False


def is_data_header_row(row_values):
    """
    Check if this row is the data header row (contains Lane #, Currency, p/unit, Applies if).
    
    The header row typically has columns like:
    - Lane # (or Lane)
    - Currency
    - p/unit (or Flat)
    - Applies if (or Apply if)
    - Valid From / Valid To (for date-based validity)
    """
    header_indicators = 0
    
    for val in row_values:
        if val is not None:
            val_str = str(val).strip().lower()
            
            # Check for various header column names
            if 'lane' in val_str:
                header_indicators += 1
            elif val_str == 'currency':
                header_indicators += 1
            elif 'p/unit' in val_str or 'per unit' in val_str:
                header_indicators += 1
            elif val_str == 'flat':
                header_indicators += 1
            elif 'apply' in val_str or 'applies' in val_str:
                header_indicators += 1
            elif 'valid from' in val_str or 'valid to' in val_str:
                header_indicators += 1
    
    # Need at least 2 header indicators to be confident it's a header row
    return header_indicators >= 2


def find_column_indices(row_values, min_max_row=None):
    """
    Find column indices for Lane #, Currency, Flat (MIN/MAX), p/unit, Apply if, Valid From/To in the data header row.
    
    Args:
        row_values: The data header row values
        min_max_row: Optional row with MIN/MAX indicators (to determine which Flat is MIN vs MAX)
    
    Returns:
        dict with keys: 'lane', 'currency', 'flat_min', 'flat_max', 'per_unit', 'apply_if', 'percentage', 'valid_from', 'valid_to'
    """
    indices = {
        'lane': None,
        'currency': None,
        'flat_min': None,
        'flat_max': None,
        'per_unit': None,
        'percentage': None,       # For % over costs (column right after Lane #)
        'apply_if': None,
        'valid_from': None,
        'valid_to': None
    }
    
    # Find all "flat" columns and their positions
    flat_positions = []
    
    for i, val in enumerate(row_values):
        if val is None:
            continue
        val_str = str(val).strip().lower()
        
        if 'lane' in val_str:
            indices['lane'] = i
        elif val_str == 'currency':
            indices['currency'] = i
        elif val_str == 'flat':
            flat_positions.append(i)
        elif 'p/unit' in val_str or 'per unit' in val_str:
            indices['per_unit'] = i
        elif 'apply if' in val_str or 'applies if' in val_str:
            indices['apply_if'] = i
        elif 'valid from' in val_str:
            indices['valid_from'] = i
        elif 'valid to' in val_str:
            indices['valid_to'] = i
    
    # Determine which flat columns are MIN vs MAX based on the min_max_row
    if min_max_row and flat_positions:
        for flat_idx in flat_positions:
            if flat_idx < len(min_max_row):
                indicator = min_max_row[flat_idx]
                if indicator is not None:
                    indicator_str = str(indicator).strip().upper()
                    if indicator_str == 'MIN':
                        indices['flat_min'] = flat_idx
                    elif indicator_str == 'MAX':
                        indices['flat_max'] = flat_idx
    elif flat_positions:
        # If no min_max_row, assume single flat column is MIN (backward compatibility)
        indices['flat_min'] = flat_positions[0]
        if len(flat_positions) > 1:
            indices['flat_max'] = flat_positions[1]
    
    return indices


def check_for_min_max_row(sheet, row_idx, max_col):
    """
    Check if the row contains MIN and/or MAX indicators.
    
    Returns:
        tuple: (has_min, has_max, row_values) where row_values can be used to map columns
    """
    row_values = get_row_values(sheet, row_idx)
    has_min = False
    has_max = False
    
    for val in row_values[:max_col]:
        if val is not None:
            val_upper = str(val).strip().upper()
            if val_upper == 'MIN':
                has_min = True
            elif val_upper == 'MAX':
                has_max = True
    
    return has_min, has_max, row_values


def extract_accessorial_costs(sheet):
    """
    Extract all accessorial cost blocks from the sheet.
    
    Structure per cost block (based on actual file):
    - Row N: Cost name (e.g., "(Prior to dispatch)")
    - Row N+1: Rate by (e.g., "Rate by: Quantity/Container")
    - Row N+2: Data header (Lane #, Currency, p/unit, Applies if)
    - Row N+3+: Actual data rows
    
    OR with MIN:
    - Row N: Cost name
    - Row N+1: Rate by
    - Row N+2: MIN row
    - Row N+3: Data header (Lane #, Currency, Flat, p/unit, Applies if)
    - Row N+4+: Actual data rows
    
    Returns:
        List of AccessorialCost objects
    """
    costs = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Start after first 2 rows (skip them)
    current_row = 3
    
    print(f"\n   Scanning rows 3 to {max_row} for cost blocks...")
    print(f"   Max columns: {max_col}")
    
    # Debug: print first rows to understand structure
    print("\n   First 25 rows preview (looking for % over costs patterns):")
    for r in range(1, min(26, max_row + 1)):
        row_vals = get_row_values(sheet, r)
        # Show ALL values with their column indices for debugging
        preview_with_idx = []
        for i, v in enumerate(row_vals[:10]):
            if v is not None:
                preview_with_idx.append(f"[{i}]:{repr(v)[:30]}")
        
        if preview_with_idx:
            # Check for special patterns in ANY cell
            all_text = ' '.join([str(v) for v in row_vals if v is not None]).lower()
            marker = ""
            if '% -' in all_text or 'over costs' in all_text:
                marker = " <-- % OVER COSTS"
            elif 'applied over' in all_text:
                marker = " <-- APPLIED OVER HEADER"
            elif any(str(v).startswith('•') for v in row_vals if v):
                marker = " <-- BULLET ITEM"
            elif 'valid from' in all_text or 'valid to' in all_text:
                marker = " <-- DATE VALIDITY"
            elif 'lane' in all_text:
                marker = " <-- LANE HEADER?"
            print(f"      Row {r}: {preview_with_idx}{marker}")
    
    while current_row <= max_row:
        row_values = get_row_values(sheet, current_row)
        
        # Get FIRST non-empty value (not necessarily column 0!)
        first_val = get_first_non_empty_value(row_values)
        first_val_idx = get_first_non_empty_index(row_values)
        
        # Debug: show raw row data for first 20 rows
        if current_row <= 20:
            print(f"   DEBUG Row {current_row}: first_val={repr(first_val)}, idx={first_val_idx}")
        
        # Skip empty rows
        if first_val is None:
            current_row += 1
            continue
        
        # Debug: check what we're evaluating
        if current_row <= 20:
            is_cost = is_cost_name_row(row_values, current_row)
            print(f"   DEBUG Row {current_row}: is_cost_name={is_cost}")
        
        if first_val and is_cost_name_row(row_values, current_row):
            # Found a cost name - extract the block
            cost_name = first_val
            print(f"\n   Found cost at row {current_row}: {cost_name}")
            
            # Look for rate_by, multiplier in subsequent rows
            multiplier = None
            rate_by = None
            apply_if = None
            has_min = False
            has_max = False
            is_percentage = False
            applied_over_costs = []
            min_max_row_values = None  # To store MIN/MAX row values for column mapping
            data_header_row = None
            
            # Scan next few rows to find the structure
            scan_row = current_row + 1
            while scan_row <= min(current_row + 15, max_row):  # Increased range for % over costs blocks
                scan_values = get_row_values(sheet, scan_row)
                # Get first non-empty value (not necessarily column 0!)
                first_scan_val = get_first_non_empty_value(scan_values)
                
                print(f"      Scanning row {scan_row}: first_val = {first_scan_val}")
                print(f"         All values: {[v for v in scan_values[:6] if v is not None]}")
                
                # Check ALL cells for percentage patterns (can be in any column)
                for cell_val in scan_values:
                    if cell_val is None:
                        continue
                    cell_str = str(cell_val).strip()
                    cell_lower = cell_str.lower()
                    
                    # Check for "% - Over costs" pattern (percentage-based cost)
                    if '% -' in cell_str or '% over' in cell_lower or 'over costs' in cell_lower:
                        is_percentage = True
                        print(f"      Row {scan_row}: PERCENTAGE-BASED COST detected: '{cell_str[:50]}'")
                        
                        # Also parse bullet points if they're in the same cell (multi-line)
                        # Split by newlines and look for bullet points
                        for line in cell_str.replace('\r\n', '\n').split('\n'):
                            line = line.strip()
                            if line.startswith('•') or (line.startswith('-') and len(line) > 2 and not line[1].isdigit()):
                                cost_item = line.lstrip('•-* ').strip()
                                if cost_item and len(cost_item) > 3:
                                    applied_over_costs.append(cost_item)
                                    print(f"      Row {scan_row}: Applied over cost (inline): {cost_item}")
                    
                    # Check for "Applied over:" pattern - may contain bullet points in same cell
                    elif cell_lower.startswith('applied over') or 'applied over' in cell_lower:
                        is_percentage = True
                        print(f"      Row {scan_row}: Applied over header found: '{cell_str[:80]}'")
                        
                        # Parse bullet points from this cell (may be multi-line)
                        for line in cell_str.replace('\r\n', '\n').split('\n'):
                            line = line.strip()
                            if line.startswith('•') or (line.startswith('-') and len(line) > 2 and not line[1].isdigit()):
                                cost_item = line.lstrip('•-* ').strip()
                                if cost_item and len(cost_item) > 3:
                                    applied_over_costs.append(cost_item)
                                    print(f"      Row {scan_row}: Applied over cost (from header cell): {cost_item}")
                    
                    # Check for standalone bullet point items (costs to apply over)
                    elif cell_str.startswith('•') or (cell_str.startswith('-') and len(cell_str) > 2 and not cell_str[1].isdigit()):
                        cost_item = cell_str.lstrip('•-* ').strip()
                        if cost_item and len(cost_item) > 3:  # Avoid false positives
                            applied_over_costs.append(cost_item)
                            print(f"      Row {scan_row}: Applied over cost: {cost_item}")
                
                if first_scan_val:
                    first_lower = first_scan_val.lower()
                    
                    # Check for rate by (e.g., "Rate by: Quantity/Container\r\nRegular rule")
                    if first_lower.startswith('rate by') or first_lower.startswith('rate:'):
                        # Extract the rate by value (after the colon if present)
                        rate_by_raw = first_scan_val
                        if ':' in rate_by_raw:
                            rate_by_raw = rate_by_raw.split(':', 1)[1].strip()
                        # Clean up newlines and "Regular rule" suffix
                        rate_by = rate_by_raw.replace('\r\n', ' ').replace('\n', ' ')
                        rate_by = rate_by.replace('Regular rule', '').strip()
                        print(f"      Row {scan_row}: Rate by = {rate_by}")
                    
                    # Check for multiplier (e.g., "per shipment", "per kg")
                    elif first_lower.startswith('per ') or 'multiplier' in first_lower:
                        multiplier = first_scan_val
                        print(f"      Row {scan_row}: Multiplier = {multiplier}")
                    
                    # Check if this is the data header row (Lane #, Currency, p/unit, Applies if)
                    elif is_data_header_row(scan_values):
                        data_header_row = scan_row
                        print(f"      Row {scan_row}: Data header row found")
                        
                        # Check row above for MIN/MAX
                        if scan_row > current_row + 1:
                            row_has_min, row_has_max, mm_row_values = check_for_min_max_row(sheet, scan_row - 1, max_col)
                            if row_has_min or row_has_max:
                                has_min = row_has_min
                                has_max = row_has_max
                                min_max_row_values = mm_row_values
                                print(f"      Row {scan_row - 1}: Has MIN={has_min}, MAX={has_max}")
                        break
                else:
                    # Check if any cell in the row indicates data header
                    if is_data_header_row(scan_values):
                        data_header_row = scan_row
                        print(f"      Row {scan_row}: Data header row found (first cell empty)")
                        break
                
                scan_row += 1
            
            # If we found a data header row, extract column indices and ALL lane data
            if data_header_row:
                header_values = get_row_values(sheet, data_header_row)
                col_indices = find_column_indices(header_values, min_max_row_values)
                print(f"      Column indices: {col_indices}")
                print(f"      Is percentage: {is_percentage}, Applied over: {applied_over_costs}")
                
                # For percentage costs, find the percentage column (usually second column after Lane #)
                # If no explicit percentage column, it's usually the column right after Lane #
                if is_percentage and col_indices['percentage'] is None:
                    # Percentage value is usually in the column immediately after Lane #
                    if col_indices['lane'] is not None:
                        col_indices['percentage'] = col_indices['lane'] + 1
                        print(f"      Percentage column auto-detected at index: {col_indices['percentage']}")
                
                # Extract ALL lane data rows
                lanes = []
                lane_row = data_header_row + 1
                
                while lane_row <= max_row:
                    lane_values = get_row_values(sheet, lane_row)
                    lane_first = get_first_non_empty_value(lane_values)
                    
                    # Stop if empty row
                    if lane_first is None:
                        break
                    
                    # Stop if it's a new cost name
                    if is_cost_name_row(lane_values, lane_row):
                        break
                    
                    # Check if it's a lane number
                    try:
                        lane_num = int(float(lane_first))
                        
                        # Extract lane data
                        currency = None
                        price_flat_min = None
                        price_flat_max = None
                        price_per_unit = None
                        percentage = None
                        applies_if = None
                        valid_from = None
                        valid_to = None
                        
                        if col_indices['currency'] is not None and col_indices['currency'] < len(lane_values):
                            currency = clean_text(lane_values[col_indices['currency']])
                        
                        # Extract percentage value (for % over costs)
                        if is_percentage and col_indices['percentage'] is not None and col_indices['percentage'] < len(lane_values):
                            pct_val = lane_values[col_indices['percentage']]
                            if pct_val is not None:
                                try:
                                    percentage = float(pct_val)
                                except (ValueError, TypeError):
                                    percentage = None
                        
                        # Extract MIN flat price (or plain flat price if no MIN/MAX indicators)
                        # Note: flat_min index is also set for plain Flat columns without MIN indicator
                        if col_indices['flat_min'] is not None and col_indices['flat_min'] < len(lane_values):
                            flat_val = lane_values[col_indices['flat_min']]
                            if flat_val is not None:
                                try:
                                    price_flat_min = float(flat_val)
                                except (ValueError, TypeError):
                                    price_flat_min = None
                        
                        # Extract MAX flat price (only if MAX indicator exists)
                        if col_indices['flat_max'] is not None and col_indices['flat_max'] < len(lane_values):
                            flat_val = lane_values[col_indices['flat_max']]
                            if flat_val is not None:
                                try:
                                    price_flat_max = float(flat_val)
                                except (ValueError, TypeError):
                                    price_flat_max = None
                        
                        if col_indices['per_unit'] is not None and col_indices['per_unit'] < len(lane_values):
                            unit_val = lane_values[col_indices['per_unit']]
                            if unit_val is not None:
                                try:
                                    price_per_unit = float(unit_val)
                                except (ValueError, TypeError):
                                    price_per_unit = None
                        
                        if col_indices['apply_if'] is not None and col_indices['apply_if'] < len(lane_values):
                            applies_if = clean_text(lane_values[col_indices['apply_if']])
                        
                        # Extract Valid From/To dates
                        if col_indices['valid_from'] is not None and col_indices['valid_from'] < len(lane_values):
                            vf_val = lane_values[col_indices['valid_from']]
                            if vf_val is not None:
                                # Handle datetime objects
                                if hasattr(vf_val, 'strftime'):
                                    valid_from = vf_val.strftime('%d.%m.%Y')
                                else:
                                    valid_from = clean_text(str(vf_val))
                        
                        if col_indices['valid_to'] is not None and col_indices['valid_to'] < len(lane_values):
                            vt_val = lane_values[col_indices['valid_to']]
                            if vt_val is not None:
                                # Handle datetime objects
                                if hasattr(vt_val, 'strftime'):
                                    valid_to = vt_val.strftime('%d.%m.%Y')
                                else:
                                    valid_to = clean_text(str(vt_val))
                        
                        # Create LaneData object
                        lane_data = LaneData(
                            lane_number=lane_num,
                            currency=currency,
                            price_flat_min=price_flat_min,
                            price_flat_max=price_flat_max,
                            price_per_unit=price_per_unit,
                            percentage=percentage,
                            applies_if=applies_if,
                            valid_from=valid_from,
                            valid_to=valid_to
                        )
                        lanes.append(lane_data)
                        
                        if is_percentage:
                            print(f"         Lane {lane_num}: {percentage}%, applies={applies_if}, valid={valid_from} to {valid_to}")
                        else:
                            print(f"         Lane {lane_num}: {currency}, flat_min={price_flat_min}, flat_max={price_flat_max}, p/unit={price_per_unit}")
                        
                        lane_row += 1
                    except (ValueError, TypeError):
                        # Not a lane number - might be end of block
                        break
                
                # Create AccessorialCost object with all lanes
                cost = AccessorialCost(
                    name=cost_name,
                    rate_by=rate_by,  # Already cleaned during scan
                    multiplier=multiplier,  # Optional, may be None
                    has_min_flat=has_min,
                    has_max_flat=has_max,
                    is_percentage=is_percentage,
                    applied_over_costs=applied_over_costs,
                    lanes=lanes,
                    start_row=current_row
                )
                costs.append(cost)
                
                if is_percentage:
                    print(f"      ✓ Cost added: {cost_name} (% over {len(applied_over_costs)} costs) with {len(lanes)} lane(s)")
                else:
                    print(f"      ✓ Cost added: {cost_name} with {len(lanes)} lane(s)")
                
                current_row = lane_row
            else:
                print(f"      [WARNING] No data header found for cost: {cost_name}")
                current_row += 1
        else:
            current_row += 1
    
    return costs


def calculate_percentage_cost(percentage_cost: AccessorialCost, 
                              base_costs: Dict[str, float], 
                              lane_number: int,
                              shipment_date: str = None) -> Optional[float]:
    """
    Calculate the actual cost for a percentage-based accessorial cost.
    
    Args:
        percentage_cost: AccessorialCost object with is_percentage=True
        base_costs: Dictionary mapping cost names to their amounts
                   e.g., {"Direct Transport (LTL)": 150.0, "Direct Transport (FTL)": 200.0}
        lane_number: The lane number to get the percentage for
        shipment_date: Optional date string (DD.MM.YYYY) to check validity
    
    Returns:
        Calculated cost amount, or None if not applicable
    
    Example:
        If Fuel Surcharge is 12% over "Direct Transport (LTL)" and "Direct Transport (FTL)",
        and base_costs = {"Direct Transport (LTL)": 150.0, "Direct Transport (FTL)": 0},
        then result = 150.0 * 0.12 = 18.0
    """
    if not percentage_cost.is_percentage:
        print(f"   [WARNING] Cost '{percentage_cost.name}' is not a percentage-based cost")
        return None
    
    # Find the lane with the given lane number
    lane_data = None
    for lane in percentage_cost.lanes:
        if lane.lane_number == lane_number:
            lane_data = lane
            break
    
    if lane_data is None:
        print(f"   [WARNING] Lane {lane_number} not found in cost '{percentage_cost.name}'")
        return None
    
    # Check date validity if provided
    if shipment_date and (lane_data.valid_from or lane_data.valid_to):
        # TODO: Add date validity check
        pass
    
    # Get percentage (e.g., 12 for 12%)
    percentage = lane_data.percentage
    if percentage is None:
        print(f"   [WARNING] No percentage value for lane {lane_number} in cost '{percentage_cost.name}'")
        return None
    
    # Sum up base costs that this percentage applies over
    total_base = 0.0
    for cost_name in percentage_cost.applied_over_costs:
        if cost_name in base_costs:
            total_base += base_costs[cost_name]
    
    # Calculate percentage cost
    calculated_cost = total_base * (percentage / 100.0)
    
    return calculated_cost


def create_costs_summary_dataframe(costs: List[AccessorialCost]) -> pd.DataFrame:
    """Create a summary DataFrame from the list of AccessorialCost objects (one row per cost)."""
    if not costs:
        return pd.DataFrame()
    
    data = [cost.to_dict() for cost in costs]
    return pd.DataFrame(data)


def create_costs_detail_dataframe(costs: List[AccessorialCost]) -> pd.DataFrame:
    """Create a detailed DataFrame with one row per lane."""
    if not costs:
        return pd.DataFrame()
    
    data = []
    for cost in costs:
        lane_dicts = cost.to_lane_dicts()
        data.extend(lane_dicts)
    
    return pd.DataFrame(data)


def extract_cost_data(sheet, cost: AccessorialCost, max_rows_per_cost=100) -> pd.DataFrame:
    """
    Extract actual data rows for a specific cost.
    
    Args:
        sheet: Worksheet object
        cost: AccessorialCost object
        max_rows_per_cost: Maximum rows to extract per cost block
    
    Returns:
        DataFrame with the cost data
    """
    if cost.data_start_row is None:
        return pd.DataFrame()
    
    data_rows = []
    current_row = cost.data_start_row
    max_row = sheet.max_row
    
    # Build header based on available columns
    header = ['Lane #']
    if cost.currency_col_idx is not None:
        header.append('Currency')
    if cost.has_min_flat and cost.price_flat_col_idx is not None:
        header.append('Price Flat MIN')
    if cost.price_per_unit_col_idx is not None:
        header.append('Price per unit')
    
    # Extract data until we hit an empty row or another cost block
    rows_extracted = 0
    while current_row <= max_row and rows_extracted < max_rows_per_cost:
        row_values = get_row_values(sheet, current_row)
        
        # Check if row is empty or starts a new cost block
        first_val = clean_text(row_values[0]) if row_values else None
        if first_val is None:
            break
        
        # Check if it's a number (Lane #) - this is data
        try:
            lane_num = float(first_val)
            # Extract relevant columns
            row_data = [lane_num]
            
            if cost.currency_col_idx is not None and cost.currency_col_idx < len(row_values):
                row_data.append(row_values[cost.currency_col_idx])
            
            if cost.has_min_flat and cost.price_flat_col_idx is not None:
                if cost.price_flat_col_idx < len(row_values):
                    row_data.append(row_values[cost.price_flat_col_idx])
            
            if cost.price_per_unit_col_idx is not None and cost.price_per_unit_col_idx < len(row_values):
                row_data.append(row_values[cost.price_per_unit_col_idx])
            
            data_rows.append(row_data)
            rows_extracted += 1
        except (ValueError, TypeError):
            # Not a lane number - might be end of block
            if is_cost_name_row(row_values, current_row):
                break
        
        current_row += 1
    
    if data_rows:
        return pd.DataFrame(data_rows, columns=header)
    return pd.DataFrame()


def save_to_excel(costs: List[AccessorialCost], sheet, agreement_number=None, output_filename=None):
    """
    Save AccessorialCost info and data to Excel file in partly_df folder.
    
    Args:
        costs: List of AccessorialCost objects
        sheet: Worksheet object (unused but kept for compatibility)
        agreement_number: Optional agreement number for filename
        output_filename: Optional explicit output filename (overrides agreement_number)
    
    Returns:
        Path to the saved file
    """
    output_folder = Path(__file__).parent / "partly_df"
    output_folder.mkdir(exist_ok=True)
    
    # Determine output filename
    if output_filename:
        filename = output_filename
    elif agreement_number:
        # Use agreement number as filename
        safe_agreement = "".join(c for c in agreement_number if c.isalnum() or c in ('-', '_', ' ')).strip()
        filename = f"{safe_agreement}_accessorial_costs.xlsx"
    else:
        filename = "accessorial_costs.xlsx"
    
    output_path = output_folder / filename
    
    # Create DataFrame with all lanes
    df_detail = create_costs_detail_dataframe(costs)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Single sheet with all lane data
            if not df_detail.empty:
                df_detail.to_excel(writer, sheet_name='Accessorial Costs', index=False)
            
        print(f"\n   Saved to: {output_path}")
        print(f"      - {len(df_detail)} lane entries from {len(costs)} cost types")
        
    except PermissionError:
        alt_filename = filename.replace('.xlsx', '_new.xlsx')
        alt_path = output_folder / alt_filename
        with pd.ExcelWriter(alt_path, engine='openpyxl') as writer:
            if not df_detail.empty:
                df_detail.to_excel(writer, sheet_name='Accessorial Costs', index=False)
        print(f"\n   [WARNING] Original file is open. Saved to: {alt_path}")
        output_path = alt_path
    
    return output_path


def process_accessorial_costs(file_path, agreement_number=None):
    """
    Process accessorial costs from a rate card file.
    
    Args:
        file_path: Path to the rate card file relative to "input/" folder
        agreement_number: Optional agreement number (if not provided, will be extracted)
    
    Returns:
        tuple: (list of AccessorialCost objects, output file path)
    """
    print(f"\n   Processing: {file_path}")
    
    # Get agreement number if not provided
    if agreement_number is None:
        agreement_number = get_agreement_number(file_path)
    
    if agreement_number:
        print(f"   Agreement number: {agreement_number}")
    else:
        print(f"   Agreement number: Not found")
    
    # Step 1: Load the Accessorial costs sheet
    print("\n   Loading Accessorial costs sheet...")
    sheet, workbook = load_accessorial_costs_sheet(file_path)
    
    # Step 2: Extract cost blocks
    print("\n   Extracting accessorial cost blocks...")
    costs = extract_accessorial_costs(sheet)
    
    print(f"\n   Total accessorial costs found: {len(costs)}")
    
    # Print summary
    if costs:
        print("\n   Cost Summary:")
        for i, cost in enumerate(costs[:5], 1):  # Show first 5
            print(f"      {i}. {cost.name} ({len(cost.lanes)} lanes)")
        if len(costs) > 5:
            print(f"      ... and {len(costs) - 5} more")
    
    # Step 3: Save to Excel
    print("\n   Saving to Excel...")
    output_path = save_to_excel(costs, sheet, agreement_number=agreement_number)
    
    # Close workbook
    workbook.close()
    
    return costs, output_path, agreement_number


def process_single_rate_card(file_path):
    """
    Process a single rate card file and save accessorial costs output.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        tuple: (output_path, agreement_number) or (None, None) if error
    """
    try:
        costs, output_path, agreement_number = process_accessorial_costs(file_path)
        return output_path, agreement_number
    except Exception as e:
        print(f"   [ERROR] Failed to process {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return None, None


def process_multiple_rate_cards(file_paths):
    """
    Process multiple rate card files and save each to a separate output file.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
                          (e.g., ["rate.xlsx", "rate_2.xlsx"])
    
    Returns:
        dict: Dictionary mapping agreement numbers to their output file paths
              {agreement_number: output_path, ...}
    """
    results = {}
    
    print(f"\n{'='*80}")
    print(f"ACCESSORIAL COSTS ANALYSIS - Processing {len(file_paths)} rate card(s)")
    print(f"{'='*80}")
    
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] {'-'*60}")
        
        output_path, agreement_number = process_single_rate_card(file_path)
        
        if output_path:
            key = agreement_number if agreement_number else os.path.splitext(os.path.basename(file_path))[0]
            results[key] = str(output_path)
    
    print(f"\n{'='*80}")
    print(f"Processing complete! {len(results)}/{len(file_paths)} files processed successfully.")
    print(f"{'='*80}")
    
    if results:
        print("\nOutput files created:")
        for agreement, path in results.items():
            print(f"  - {agreement}: {path}")
    
    return results


def get_rate_card_files_from_input():
    """
    Get all Excel files from the input folder that could be rate cards.
    
    Returns:
        list: List of Excel file names in the input folder
    """
    input_folder = "input"
    if not os.path.exists(input_folder):
        print(f"Warning: Input folder '{input_folder}' does not exist.")
        return []
    
    excel_files = [f for f in os.listdir(input_folder) 
                   if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    return excel_files


def main():
    """Main function to run the accessorial costs analysis for multiple rate cards."""
    print("\n" + "="*80)
    print("ACCESSORIAL COSTS ANALYSIS")
    print("="*80)
    
    # Process multiple rate cards (same as part4_rate_card_processing.py)
    rate_card_files = ["ra_densir.xlsx"]
    results = process_multiple_rate_cards(rate_card_files)
    
    return results


if __name__ == "__main__":
    results = main()

