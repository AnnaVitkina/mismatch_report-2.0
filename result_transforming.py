"""
Result Transforming - Excel Formatting

This module handles the final formatting of result Excel files:
- Header styling
- Alternating row colors for cost type groups
- Column width adjustment
- Borders and alignment
- Freeze panes
- Add extra columns from source files
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Column renaming mapping (original -> new name) - ALWAYS APPLIED
COLUMN_RENAME_MAP = {
    'ETOF_NUMBER': 'ETOF',
    'SHIPMENT_ID': 'Shipment ID',
    'DELIVERY_NUMBER': 'Delivery Number',
    'SHIP_DATE': 'Shipment date',
    'SHIP_COUNTRY_ETOF': 'Origin country',
    'SHIP_CITY_ETOF': 'Origin city',
    'CUST_COUNTRY_ETOF': 'Destination country',
    'CUST_CITY_ETOF': 'Destination city',
    'SERVICE_ETOF': 'Service',
}


# =============================================================================
# EXTRA COLUMNS ALIAS MAP
# =============================================================================
# This mapping defines what the user can type -> what actual column name to look for
# Format: 'what user can type': 'actual column name in source file'
# Add your own mappings here!
# =============================================================================
EXTRA_COLUMNS_ALIAS_MAP = {
    # Format: 'Display name (from dropdown)': 'Actual column name in source file'
    'Invoice entity': 'INVOICE_ENTITY',
    'Carrier name': 'CARRIER_NAME',
    'Destination postal code': 'CUST_POST',
    'Origin postal code': 'SHIP_POST',
    'Destination airport': 'CUST_AIRPORT',
    'Equipment type': 'CONT_LOAD',
    'Origin airport': 'SHIP_AIRPORT',
    'Business unit name': 'BU_NAME',
    'Transport mode': 'TRANSPORT_MODE',
    'LDM': 'LDM',
    'CBM': 'CBM',
    'Weight': 'WEIGHT',
    'DANGEROUS Goods': 'DANGEROUS_GOODS',
    'Charge weight': 'CHARGE_WEIGHT',
    'House bill': 'HOUSE_BILL',
    'Master bill': 'MASTER_BILL',
    'Roundtrip': 'ROUNDTRIP',
}


def get_column_aliases(col_name):
    """
    Get all possible column names to search for based on user input.
    
    Checks EXTRA_COLUMNS_ALIAS_MAP first, then returns as-is if not found.
    
    Args:
        col_name: Column name provided by user
    
    Returns:
        List of possible column names to search for in source file
    """
    col_stripped = col_name.strip()
    col_lower = col_stripped.lower()
    
    # Check if user input matches any alias in EXTRA_COLUMNS_ALIAS_MAP
    for alias, actual_col in EXTRA_COLUMNS_ALIAS_MAP.items():
        if col_lower == alias.lower():
            # Return both the actual column and the alias (in case source uses alias)
            return [actual_col, col_stripped]
    
    # If not in alias map, return as-is
    return [col_stripped]


def rename_columns(df):
    """
    Rename columns according to the COLUMN_RENAME_MAP.
    
    Args:
        df: DataFrame to rename columns in
    
    Returns:
        DataFrame with renamed columns
    """
    # Create a case-insensitive mapping
    rename_map = {}
    for col in df.columns:
        col_upper = col.upper().strip()
        for old_name, new_name in COLUMN_RENAME_MAP.items():
            if col_upper == old_name.upper():
                rename_map[col] = new_name
                break
    
    if rename_map:
        print(f"      Renaming columns: {rename_map}")
        df = df.rename(columns=rename_map)
    
    return df


def get_partly_df_folder():
    """Get the path to the partly_df folder."""
    return Path(__file__).parent / "partly_df"


def find_etof_column(df):
    """Find the ETOF number column in a DataFrame."""
    etof_patterns = ['etof', 'etof_number', 'etof number', 'etof#', 'etof #']
    for col in df.columns:
        col_lower = col.lower().strip()
        for pattern in etof_patterns:
            if pattern in col_lower:
                return col
    return None


def add_columns_from_source(result_df, columns_to_add, sheet_name=None):
    """
    Add specified columns from lc_etof_with_comments.xlsx to the result DataFrame.
    
    Args:
        result_df: DataFrame to add columns to
        columns_to_add: List of column names to extract and add
        sheet_name: Optional sheet name for logging
    
    Returns:
        DataFrame with added columns
    """
    if not columns_to_add:
        return result_df
    
    partly_df = get_partly_df_folder()
    source_file = partly_df / "lc_etof_with_comments.xlsx"
    
    if not source_file.exists():
        print(f"      [WARNING] Source file not found: {source_file}")
        return result_df
    
    # Load source file
    try:
        source_df = pd.read_excel(source_file)
        print(f"      Loaded source file: {len(source_df)} rows")
    except Exception as e:
        print(f"      [WARNING] Error loading source file: {e}")
        return result_df
    
    # Find ETOF column in both DataFrames
    result_etof_col = find_etof_column(result_df)
    source_etof_col = find_etof_column(source_df)
    
    if result_etof_col is None:
        print(f"      [WARNING] ETOF column not found in result DataFrame")
        return result_df
    
    if source_etof_col is None:
        print(f"      [WARNING] ETOF column not found in source file")
        return result_df
    
    print(f"      Matching on: result[{result_etof_col}] <-> source[{source_etof_col}]")
    
    # Find which columns exist in source (using alias mapping)
    columns_found = []
    columns_not_found = []
    
    for col in columns_to_add:
        col_stripped = col.strip()
        found = False
        
        # Get possible aliases from EXTRA_COLUMNS_ALIAS_MAP
        aliases = get_column_aliases(col_stripped)
        
        # Try to find any of the aliases in the source DataFrame
        for alias in aliases:
            matching_cols = [c for c in source_df.columns if c.lower().strip() == alias.lower()]
            if matching_cols:
                columns_found.append(matching_cols[0])  # Use the actual column name from source
                if alias != col_stripped:
                    print(f"      Column '{col_stripped}' mapped to '{matching_cols[0]}'")
                found = True
                break
        
        if not found:
            columns_not_found.append(col_stripped)
    
    if columns_not_found:
        print(f"      [WARNING] Columns not found in source: {columns_not_found}")
        print(f"      Tip: Add mappings to EXTRA_COLUMNS_ALIAS_MAP in result_transforming.py")
    
    if not columns_found:
        print(f"      [WARNING] No requested columns found in source file")
        return result_df
    
    print(f"      Adding columns: {columns_found}")
    
    # Create a subset of source with ETOF and requested columns
    source_subset = source_df[[source_etof_col] + columns_found].copy()
    
    # Remove duplicates based on ETOF (keep first)
    source_subset = source_subset.drop_duplicates(subset=[source_etof_col], keep='first')
    
    # Rename source ETOF column to match result for merging
    source_subset = source_subset.rename(columns={source_etof_col: result_etof_col})
    
    # Merge with result
    result_with_cols = result_df.merge(
        source_subset,
        on=result_etof_col,
        how='left',
        suffixes=('', '_added')
    )
    
    print(f"      Columns added successfully")
    return result_with_cols


def apply_formatting(wb, cost_type_groups=None):
    """
    Apply formatting to the workbook.
    
    Args:
        wb: openpyxl Workbook object
        cost_type_groups: dict {sheet_name: list of (start_row, end_row, color_index), ...}
                         where color_index is 0 or 1 for alternating colors
    """
    if cost_type_groups is None:
        cost_type_groups = {}
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    pivot_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Two alternating colors for cost type groups
    cost_color_1 = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')  # Light blue
    cost_color_2 = None  # White (no fill)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Determine if this is a pivot sheet
        is_pivot = 'Pivot' in sheet_name
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = pivot_header_fill if is_pivot else header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Get cost type groups for this sheet
        groups = cost_type_groups.get(sheet_name, [])
        
        # Create a row -> color mapping from groups
        row_colors = {}
        for start_row, end_row, color_idx in groups:
            for r in range(start_row, end_row + 1):
                row_colors[r] = color_idx
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # Apply color based on cost type group (for data sheets)
                if not is_pivot and row_idx in row_colors:
                    if row_colors[row_idx] == 0:
                        cell.fill = cost_color_1
                    # else: leave white (no fill needed)
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'


def format_result_file(file_path, cost_type_groups=None, extra_columns=None):
    """
    Load an Excel file, rename columns, add extra columns, apply formatting, and save it.
    
    Args:
        file_path: Path to the Excel file to format
        cost_type_groups: dict {sheet_name: list of (start_row, end_row, color_index), ...}
        extra_columns: List of column names to add from lc_etof_with_comments.xlsx
    
    Returns:
        Path to the formatted file
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Step 1: Load all sheets, rename columns, and optionally add extra columns
    print(f"\n   Transforming result file: {file_path}")
    
    xlsx = pd.ExcelFile(file_path)
    transformed_sheets = {}
    
    for sheet_name in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet_name)
        
        # Only process data sheets (not pivot sheets)
        if 'Pivot' not in sheet_name:
            print(f"\n   Processing sheet: {sheet_name}")
            
            # Rename columns
            df = rename_columns(df)
            
            # Add extra columns if specified
            if extra_columns:
                print(f"      Adding extra columns: {extra_columns}")
                df = add_columns_from_source(df, extra_columns, sheet_name)
        
        transformed_sheets[sheet_name] = df
    
    # Save back to file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, df in transformed_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"\n   Columns renamed and saved")
    
    # Step 2: Apply formatting
    print(f"\n   Applying formatting to: {file_path}")
    
    wb = load_workbook(file_path)
    apply_formatting(wb, cost_type_groups)
    wb.save(file_path)
    
    print(f"   Formatting applied successfully")
    return file_path


def main(file_path=None, cost_type_groups=None, extra_columns=None):
    """
    Main function to format a result file.
    
    Args:
        file_path: Path to the file to format. If None, uses default output/result.xlsx
        cost_type_groups: dict {sheet_name: list of (start_row, end_row, color_index), ...}
        extra_columns: List of column names to add from lc_etof_with_comments.xlsx
    
    Returns:
        Path to the formatted file
    """
    if file_path is None:
        file_path = Path(__file__).parent / "output" / "result.xlsx"
    
    return format_result_file(file_path, cost_type_groups, extra_columns)


if __name__ == "__main__":
    main()
