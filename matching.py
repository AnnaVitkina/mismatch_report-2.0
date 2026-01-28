"""
Rate Lane Matching Script

This script:
1. Gets rate cards from partly_df/ folder (created by part4_rate_card_processing.py)
2. Gets vocabulary mapping files from partly_df/ folder (created by vocabular.py)
3. For each agreement, matches shipments with rate card to find Lane #
4. Creates separate output files per agreement
5. Validates business rules (postal code zones, country regions) to find exact matches
"""

import pandas as pd
import os
import re
import sys
from pathlib import Path
from datetime import datetime, date


class OutputLogger:
    """Logger that writes output to both console and a text file."""
    
    def __init__(self, log_file_path):
        self.log_file_path = log_file_path
        self.log_file = None
        self.original_stdout = sys.stdout
        
    def __enter__(self):
        # Open log file for writing
        self.log_file = open(self.log_file_path, 'w', encoding='utf-8')
        # Write header with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_file.write(f"Matching Log - Generated: {timestamp}\n")
        self.log_file.write("=" * 80 + "\n\n")
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.log_file:
            self.log_file.close()
        return False
    
    def write(self, message):
        """Write message to both console and file."""
        self.original_stdout.write(message)
        if self.log_file:
            self.log_file.write(message)
            self.log_file.flush()  # Ensure immediate write
    
    def flush(self):
        self.original_stdout.flush()
        if self.log_file:
            self.log_file.flush()


# Global logger instance
_output_logger = None


def init_logger(log_file_path=None):
    """Initialize the output logger."""
    global _output_logger
    if log_file_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file_path = f"partly_df/matching_output_{timestamp}.txt"
    _output_logger = OutputLogger(log_file_path)
    return _output_logger


def log_print(*args, **kwargs):
    """Print function that also logs to file if logger is active."""
    message = ' '.join(str(arg) for arg in args)
    end = kwargs.get('end', '\n')
    
    # Print to console
    print(*args, **kwargs)
    
    # Write to log file if logger is active
    global _output_logger
    if _output_logger and _output_logger.log_file:
        _output_logger.log_file.write(message + end)
        _output_logger.log_file.flush()

# Import business rules functions for validating postal code zones and country regions
try:
    from part4_rate_card_processing import (
        get_business_rules_lookup,
        process_business_rules,
        transform_business_rules_to_conditions,
        find_business_rule_columns
    )
    BUSINESS_RULES_AVAILABLE = True
except ImportError:
    print("[WARNING] Could not import business rules functions from part4_rate_card_processing.py")
    BUSINESS_RULES_AVAILABLE = False

def normalize_value(value):
    """Converts a value to lowercase string, removes spaces and underscores, and handles NaN.
    Preserves leading zeros for postal codes and similar values."""
    if pd.isna(value):
        return None

    # Convert to string first
    str_value = str(value).strip()
    
    # Check if value starts with '0' and has more digits - likely a postal code or code with leading zeros
    # Don't convert to number to preserve leading zeros (e.g., "04123" should stay "04123", not become "4123")
    if str_value.startswith('0') and len(str_value) > 1 and str_value.lstrip('0').isdigit():
        # Preserve as string to keep leading zeros
        pass
    else:
        # Attempt to convert to a number if it looks like one, then convert to int if possible
        try:
            # Convert to float for numeric conversion
            num_val = float(str_value)
            if num_val == int(num_val):  # Check if it's an integer number (e.g., 7719.0)
                str_value = str(int(num_val))
            else:  # Keep as float if it has decimal (e.g., 123.45)
                str_value = str(num_val)
        except (ValueError, TypeError):
            # Not a number, keep original value
            pass

    # Apply lowercasing and cleaning
    return str_value.lower().replace(" ", "").replace("_", "")


def normalize_column_name(col_name):
    """Normalize column names for comparison (lowercase, remove spaces/underscores)."""
    if col_name is None:
        return None
    return str(col_name).lower().replace(" ", "").replace("_", "")


def parse_date_flexible(date_value):
    """
    Parse date from various formats and return a date object.
    
    Supports formats:
    - YYYYMMDD (e.g., 20251214)
    - DD.MM.YYYY (e.g., 14.12.2025)
    - YYYY-MM-DD (e.g., 2025-12-14)
    - Excel datetime objects
    
    Args:
        date_value: Date value in various formats
        
    Returns:
        date object or None if parsing fails
    """
    if pd.isna(date_value) or date_value is None:
        return None
    
    # If already a datetime or date object
    if isinstance(date_value, (datetime, date)):
        return date_value if isinstance(date_value, date) else date_value.date()
    
    date_str = str(date_value).strip()
    if not date_str or date_str.lower() == 'nan':
        return None
    
    # Try YYYYMMDD format (20251214)
    if len(date_str) == 8 and date_str.isdigit():
        try:
            year = int(date_str[0:4])
            month = int(date_str[4:6])
            day = int(date_str[6:8])
            return date(year, month, day)
        except (ValueError, IndexError):
            pass
    
    # Try DD.MM.YYYY format (14.12.2025)
    if '.' in date_str:
        parts = date_str.split('.')
        if len(parts) == 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                return date(year, month, day)
            except (ValueError, IndexError):
                pass
    
    # Try YYYY-MM-DD format (2025-12-14)
    if '-' in date_str:
        parts = date_str.split('-')
        if len(parts) == 3:
            try:
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return date(year, month, day)
            except (ValueError, IndexError):
                pass
    
    # Try DD/MM/YYYY format (14/12/2025)
    if '/' in date_str:
        parts = date_str.split('/')
        if len(parts) == 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                return date(year, month, day)
            except (ValueError, IndexError):
                pass
    
    return None


def is_date_in_validity_period(ship_date, valid_from, valid_to, debug=False):
    """
    Check if a shipment date falls within the validity period of a lane.
    
    Args:
        ship_date: Shipment date (any format supported by parse_date_flexible)
        valid_from: Valid From date (any format)
        valid_to: Valid To date (any format)
        debug: Enable debug output
        
    Returns:
        tuple: (is_valid: bool or None, reason: str)
               - True: date is within validity period
               - False: date is outside validity period
               - None: no validity dates specified (lane is always valid)
    """
    ship_date_obj = parse_date_flexible(ship_date)
    valid_from_obj = parse_date_flexible(valid_from)
    valid_to_obj = parse_date_flexible(valid_to)
    
    # If no validity dates specified, lane is always valid
    if valid_from_obj is None and valid_to_obj is None:
        return None, "No validity period specified"
    
    # If no ship date, we can't validate
    if ship_date_obj is None:
        if debug:
            print(f"      [DATE] Cannot validate - ship_date is missing or invalid: '{ship_date}'")
        return None, "Ship date missing or invalid"
    
    # Check Valid From
    if valid_from_obj is not None:
        if ship_date_obj < valid_from_obj:
            reason = f"Ship date {ship_date_obj} is before Valid From {valid_from_obj}"
            if debug:
                print(f"      [DATE] DISQUALIFIED: {reason}")
            return False, reason
    
    # Check Valid To
    if valid_to_obj is not None:
        if ship_date_obj > valid_to_obj:
            reason = f"Ship date {ship_date_obj} is after Valid To {valid_to_obj}"
            if debug:
                print(f"      [DATE] DISQUALIFIED: {reason}")
            return False, reason
    
    # Date is within validity period
    reason = f"Ship date {ship_date_obj} is within validity period"
    if valid_from_obj and valid_to_obj:
        reason += f" ({valid_from_obj} to {valid_to_obj})"
    elif valid_from_obj:
        reason += f" (from {valid_from_obj})"
    elif valid_to_obj:
        reason += f" (until {valid_to_obj})"
    
    if debug:
        print(f"      [DATE] VALID: {reason}")
    
    return True, reason


# Note: extract_country_code is already applied in part1_etof_file_processing.py
# No need to duplicate it here as vocabular.py uses processed dataframes


def load_business_rules_for_matching(rate_card_file_path):
    """
    Load business rules and create lookup structures for matching.
    
    Args:
        rate_card_file_path: Path to rate card file (relative to input folder)
    
    Returns:
        dict: Business rules lookup with:
            - 'rule_to_country': {rule_name: country_code}
            - 'rule_to_postal_codes': {rule_name: [postal_codes]}
            - 'business_rule_columns': set of column names containing business rules
            - 'column_to_rules': {column_name: [rule_names found in it]}
    """
    if not BUSINESS_RULES_AVAILABLE:
        return None
    
    try:
        result = get_business_rules_lookup(rate_card_file_path)
        if result:
            print(f"   [Business Rules] Loaded {len(result.get('rule_to_country', {}))} rules with country mappings")
            print(f"   [Business Rules] Loaded {len(result.get('rule_to_postal_codes', {}))} rules with postal codes")
            excluded_rules = result.get('rule_to_excluded_postal_codes', {})
            print(f"   [Business Rules] Loaded {len(excluded_rules)} rules with EXCLUDED postal codes")
            if excluded_rules:
                for rule_name, codes in list(excluded_rules.items())[:5]:
                    print(f"      - {rule_name}: {len(codes)} excluded codes")
        return result
    except Exception as e:
        print(f"   [WARNING] Could not load business rules: {e}")
        return None


def validate_business_rule_value(rule_name, shipment_row, business_rules_lookup, column_name, debug=False):
    """
    Validate if shipment data matches a business rule (e.g., Zone C, HUB MADRID).
    
    Args:
        rule_name: The business rule name from rate card (e.g., "Zone C", "HUB MADRID")
        shipment_row: Row from shipment dataframe
        business_rules_lookup: Business rules lookup dictionary
        column_name: Name of the column being validated (to determine origin vs destination)
        debug: Enable debug output
    
    Returns:
        tuple: (is_valid, reason)
            - is_valid: True if shipment matches the business rule
            - reason: Explanation string
    """
    if business_rules_lookup is None or rule_name is None:
        return None, "No business rules available"
    
    rule_name_str = str(rule_name).strip()
    if not rule_name_str or rule_name_str.lower() in ['nan', 'none', '']:
        return None, "Empty rule name"
    
    rule_to_country = business_rules_lookup.get('rule_to_country', {})
    rule_to_postal = business_rules_lookup.get('rule_to_postal_codes', {})
    
    if debug:
        print(f"      [BR DEBUG] Validating rule '{rule_name_str}' for column '{column_name}'")
        print(f"      [BR DEBUG] Available rules with country: {list(rule_to_country.keys())[:10]}...")
        print(f"      [BR DEBUG] Available rules with postal: {list(rule_to_postal.keys())[:10]}...")
    
    # Find the rule (case-insensitive matching)
    matched_rule = None
    for name in rule_to_country.keys():
        if str(name).lower().strip() == rule_name_str.lower():
            matched_rule = name
            break
    
    # Also check in postal codes if not found
    if not matched_rule:
        for name in rule_to_postal.keys():
            if str(name).lower().strip() == rule_name_str.lower():
                matched_rule = name
                break
    
    if not matched_rule:
        if debug:
            print(f"      [BR DEBUG] Rule '{rule_name_str}' NOT FOUND in business rules!")
        return None, f"Rule '{rule_name_str}' not found in business rules"
    
    expected_country = rule_to_country.get(matched_rule)
    expected_postal_codes = rule_to_postal.get(matched_rule, [])
    
    # Get excluded postal codes from lookup
    rule_to_excluded = business_rules_lookup.get('rule_to_excluded_postal_codes', {})
    excluded_postal_codes = rule_to_excluded.get(matched_rule, [])
    
    # If no exclusions in lookup, try to load directly from Excel file
    if not excluded_postal_codes and not rule_to_excluded:
        try:
            # Try to find and load exclusions from rate card Excel
            import glob
            excel_files = glob.glob("partly_df/RA*.xlsx")
            for excel_file in excel_files:
                if '~$' in excel_file:  # Skip temp files
                    continue
                try:
                    xl = pd.ExcelFile(excel_file)
                    if 'Business Rules' in xl.sheet_names:
                        df_br = pd.read_excel(xl, sheet_name='Business Rules')
                        if 'Excluded Postal Codes' in df_br.columns and 'Rule Name' in df_br.columns:
                            # Find this specific rule
                            rule_row = df_br[df_br['Rule Name'] == matched_rule]
                            if not rule_row.empty:
                                excluded_str = rule_row.iloc[0].get('Excluded Postal Codes', '')
                                if pd.notna(excluded_str) and str(excluded_str).strip():
                                    excluded_postal_codes = [c.strip() for c in str(excluded_str).split(',') if c.strip()]
                                    if debug:
                                        print(f"      [BR DEBUG] Loaded exclusions directly from {excel_file}")
                            break
                except Exception as e:
                    if debug:
                        print(f"      [BR DEBUG] Could not read {excel_file}: {e}")
        except Exception as e:
            if debug:
                print(f"      [BR DEBUG] Error loading exclusions: {e}")
    
    if debug:
        print(f"      [BR DEBUG] Matched rule: '{matched_rule}'")
        print(f"      [BR DEBUG] Expected country: {expected_country}")
        print(f"      [BR DEBUG] Expected postal codes: {expected_postal_codes}")
        # Always show excluded info for debugging
        print(f"      [BR DEBUG] Total rules with exclusions in lookup: {len(rule_to_excluded)}")
        if excluded_postal_codes:
            print(f"      [BR DEBUG] EXCLUDED postal codes for '{matched_rule}': {excluded_postal_codes[:20]}{'...' if len(excluded_postal_codes) > 20 else ''}")
        else:
            print(f"      [BR DEBUG] NO excluded postal codes found for '{matched_rule}'")
    
    # Determine if this is origin or destination based on column name
    col_lower = str(column_name).lower()
    is_origin = 'origin' in col_lower or 'ship' in col_lower or 'from' in col_lower
    is_destination = 'destination' in col_lower or 'cust' in col_lower or 'to' in col_lower
    
    # Determine if this is a Country Region rule (only country matters) or Postal Code Zone (country + postal)
    is_country_region = 'country' in col_lower and 'region' in col_lower
    
    if debug:
        print(f"      [BR DEBUG] Column type: is_origin={is_origin}, is_destination={is_destination}, is_country_region={is_country_region}")
    
    # Get actual country and postal from shipment
    actual_country = None
    actual_postal = None
    country_col_found = None
    postal_col_found = None
    
    if is_origin:
        country_cols = ['Origin Country', 'origin country', 'OriginCountry', 'ORIGIN_COUNTRY']
        postal_cols = ['Origin Postal Code', 'origin postal code', 'OriginPostalCode', 'ORIGIN_POSTAL_CODE', 
                      'Ship Postal', 'SHIP_POST', 'Ship_Post']
    elif is_destination:
        country_cols = ['Destination Country', 'destination country', 'DestinationCountry', 'DESTINATION_COUNTRY',
                       'Cust Country', 'CUST_COUNTRY']
        postal_cols = ['Destination Postal Code', 'destination postal code', 'DestinationPostalCode',
                      'DESTINATION_POSTAL_CODE', 'Cust Postal', 'CUST_POST', 'Cust_Post']
    else:
        return None, f"Cannot determine origin/destination from column '{column_name}'"
    
    # Find actual values from shipment
    for col in country_cols:
        if col in shipment_row.index:
            val = shipment_row.get(col)
            if pd.notna(val) and str(val).strip():
                actual_country = str(val).strip().upper()
                country_col_found = col
                break
    
    for col in postal_cols:
        if col in shipment_row.index:
            val = shipment_row.get(col)
            if pd.notna(val) and str(val).strip():
                actual_postal = str(val).strip()
                postal_col_found = col
                # Handle postal codes with "/" (e.g., "46001/18c20250917" -> "46001")
                if '/' in actual_postal:
                    actual_postal = actual_postal.split('/')[0]
                break
    
    if debug:
        print(f"      [BR DEBUG] Shipment country ({country_col_found}): '{actual_country}'")
        print(f"      [BR DEBUG] Shipment postal ({postal_col_found}): '{actual_postal}'")
    
    # Validate country
    country_valid = True
    if expected_country:
        if actual_country:
            # Handle multiple countries (comma-separated)
            expected_countries = [c.strip().upper() for c in str(expected_country).split(',') if c.strip()]
            country_valid = actual_country in expected_countries
            if debug:
                print(f"      [BR DEBUG] Country check: '{actual_country}' in {expected_countries} = {country_valid}")
        else:
            country_valid = False
            if debug:
                print(f"      [BR DEBUG] Country check: FAILED (no actual country)")
    
    if not country_valid:
        return False, f"Country mismatch: shipment has '{actual_country}', rule expects '{expected_country}'"
    
    # Validate postal code (skip for country-only region rules)
    postal_valid = True
    is_excluded = False
    matched_exclude_prefix = None
    
    if not is_country_region and expected_postal_codes:
        if actual_postal:
            # Check if actual postal starts with any expected postal code prefix
            postal_valid = False
            actual_postal_lower = actual_postal.lower()
            for expected_pc in expected_postal_codes:
                expected_pc_str = str(expected_pc).strip().lower()
                if actual_postal_lower.startswith(expected_pc_str) or actual_postal_lower == expected_pc_str:
                    postal_valid = True
                    if debug:
                        print(f"      [BR DEBUG] Postal check: '{actual_postal}' starts with '{expected_pc_str}' = MATCH")
                    break
            if not postal_valid and debug:
                print(f"      [BR DEBUG] Postal check: '{actual_postal}' NOT starts with any of {expected_postal_codes} = FAIL")
            
            # Check excluded postal codes (only if postal was initially valid)
            if postal_valid and excluded_postal_codes:
                for excluded_pc in excluded_postal_codes:
                    excluded_pc_str = str(excluded_pc).strip().lower()
                    if actual_postal_lower.startswith(excluded_pc_str) or actual_postal_lower == excluded_pc_str:
                        is_excluded = True
                        matched_exclude_prefix = excluded_pc_str
                        if debug:
                            print(f"      [BR DEBUG] EXCLUSION check: '{actual_postal}' starts with EXCLUDED '{excluded_pc_str}' = EXCLUDED!")
                        break
                if not is_excluded and debug:
                    print(f"      [BR DEBUG] EXCLUSION check: '{actual_postal}' NOT in excluded list = OK")
        else:
            postal_valid = False
            if debug:
                print(f"      [BR DEBUG] Postal check: FAILED (no actual postal code)")
    elif is_country_region and debug:
        print(f"      [BR DEBUG] Postal check: SKIPPED (country region rule)")
    
    if not postal_valid:
        return False, f"Postal code mismatch: shipment has '{actual_postal}', rule expects prefix from {expected_postal_codes}"
    
    if is_excluded:
        return False, f"Postal code EXCLUDED: shipment has '{actual_postal}', which starts with excluded prefix '{matched_exclude_prefix}'"
    
    if debug:
        print(f"      [BR DEBUG] VALIDATION PASSED!")
    
    return True, f"Business rule '{matched_rule}' validated: Country={expected_country}, Postal={expected_postal_codes}"


def load_conditions():
    """Load conditional rules from Filtered_Rate_Card_with_Conditions.xlsx."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    conditions_file = os.path.join(script_dir, "Filtered_Rate_Card_with_Conditions.xlsx")
    
    if not os.path.exists(conditions_file):
        print(f"Warning: {conditions_file} not found. Conditions will not be validated.")
        return {}
    
    try:
        df_conditions = pd.read_excel(conditions_file, sheet_name='Conditions')
        
        # Parse conditions into a dictionary: {column_name: [list of condition rules]}
        conditions_dict = {}
        current_column = None
        
        for _, row in df_conditions.iterrows():
            column = row.get('Column', '')
            condition_rule = row.get('Condition Rule', '')
            
            if pd.notna(column) and str(column).strip() and str(column).strip() != 'nan':
                current_column = str(column).strip()
                if current_column not in conditions_dict:
                    conditions_dict[current_column] = []
            
            if pd.notna(condition_rule) and str(condition_rule).strip() and current_column:
                condition_text = str(condition_rule).strip()
                # Skip header lines like "Conditional rules:"
                if condition_text.lower() not in ['conditional rules:', 'conditional rules']:
                    conditions_dict[current_column].append(condition_text)
        
        print(f"\nLoaded conditions for {len(conditions_dict)} columns")
        return conditions_dict
    except Exception as e:
        print(f"Warning: Could not load conditions: {e}")
        return {}


def parse_condition(condition_text, rate_card_value):
    """Parse a condition rule and extract the value it applies to.
    
    Example: "NAC: RATE_TYPE is empty in any item and does not contain FAK in any item"
    Returns: ('NAC', condition_logic)
    """
    if not condition_text or pd.isna(condition_text):
        return None, None
    
    condition_text = str(condition_text).strip()
    
    # Check if condition starts with a value followed by colon (e.g., "NAC: ...")
    if ':' in condition_text:
        parts = condition_text.split(':', 1)
        condition_value = parts[0].strip()
        condition_logic = parts[1].strip() if len(parts) > 1 else ''
        
        return condition_value, condition_logic
    
    return None, condition_text


def value_satisfies_condition(resmed_value, rate_card_value, condition_text):
    """Check if a ResMed value satisfies the condition for a given rate card value.
    
    Args:
        resmed_value: The value from ResMed dataframe
        rate_card_value: The value from Rate Card (e.g., 'NAC')
        condition_text: The condition rule text
    
    Returns:
        True if the value satisfies the condition, False otherwise
    
    Example:
        condition_text = "NAC: RATE_TYPE is empty in any item and does not contain FAK in any item"
        rate_card_value = "NAC"
        resmed_value = nan (empty)
        Returns: True (because empty satisfies "is empty")
    """
    if not condition_text or pd.isna(condition_text):
        return False
    
    condition_text = str(condition_text).strip()
    condition_lower = condition_text.lower()
    rate_card_val_str = str(rate_card_value).lower() if pd.notna(rate_card_value) else ''
    
    # Check if condition is for this rate card value (format: "1. NAC: ..." or "NAC: ...")
    if ':' in condition_text:
        # Handle numbered conditions like "1. NAC:" or "1.NAC:" or just "NAC:"
        # Remove leading number and dot if present (e.g., "1. " or "1.")
        condition_text_cleaned = re.sub(r'^\d+\.\s*', '', condition_text)
        condition_parts = condition_text_cleaned.split(':', 1)
        condition_value = condition_parts[0].strip()
        condition_logic = condition_parts[1].strip() if len(condition_parts) > 1 else ''
        
        # Check if this condition applies to the rate card value
        if rate_card_val_str and condition_value.lower() != rate_card_val_str:
            return False
        
        condition_text = condition_logic  # Use only the logic part
        condition_lower = condition_text.lower()
    
    # Check if ResMed value is empty/NaN
    is_empty = pd.isna(resmed_value) or str(resmed_value).strip() == '' or str(resmed_value).lower() in ['nan', 'none', 'null', '']
    resmed_val_str = str(resmed_value).lower() if pd.notna(resmed_value) else ''
    
    # Parse condition logic
    # Example: "RATE_TYPE is empty in any item and does not contain FAK in any item"
    
    # Check "is empty" condition
    if 'is empty' in condition_lower or 'is empty in any item' in condition_lower:
        if is_empty:
            # Value is empty - check if there are additional conditions
            # If condition has "and does not contain", empty values satisfy this (empty doesn't contain anything)
            if 'does not contain' in condition_lower or 'and' in condition_lower:
                # For "and" conditions, all must be satisfied
                # Empty value satisfies "is empty" and "does not contain X" (empty doesn't contain anything)
                return True
            return True
    
    # Check "does not contain" condition
    if 'does not contain' in condition_lower:
        if is_empty:
            return True  # Empty values don't contain anything
        
        # Extract what it should not contain
        parts = condition_lower.split('does not contain')
        if len(parts) > 1:
            forbidden_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values (e.g., "EY,ETIHAD,ETIHAD AIRWAYS")
            forbidden_values = [v.strip() for v in forbidden_part.split(',')]
            # Check if ResMed value contains any forbidden value
            for forbidden in forbidden_values:
                if forbidden and forbidden in resmed_val_str:
                    return False  # Contains forbidden value - condition not satisfied
            return True  # Doesn't contain any forbidden value
    
    # Check "does not equal" condition
    if 'does not equal' in condition_lower or 'does not equal to' in condition_lower:
        if is_empty:
            return True  # Empty values don't equal anything
        
        parts = condition_lower.split('does not equal')
        if len(parts) > 1:
            forbidden_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values
            forbidden_values = [v.strip() for v in forbidden_part.split(',')]
            # Check if ResMed value equals any forbidden value
            for forbidden in forbidden_values:
                if forbidden and resmed_val_str == forbidden:
                    return False  # Equals forbidden value - condition not satisfied
            return True  # Doesn't equal any forbidden value
    
    # Check "contains" condition (positive match)
    if 'contains' in condition_lower and 'does not contain' not in condition_lower:
        if is_empty:
            return False  # Empty values don't contain anything
        
        parts = condition_lower.split('contains')
        if len(parts) > 1:
            required_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values
            required_values = [v.strip() for v in required_part.split(',')]
            # Check if ResMed value contains any required value
            for required in required_values:
                if required and required in resmed_val_str:
                    return True  # Contains required value
            return False  # Doesn't contain any required value
    
    # Check "equals" or "equal to" condition
    if 'equal to' in condition_lower or ('equals' in condition_lower and 'does not equal' not in condition_lower):
        if is_empty:
            return False  # Empty values don't equal anything
        
        if 'equal to' in condition_lower:
            parts = condition_lower.split('equal to')
        else:
            parts = condition_lower.split('equals')
        if len(parts) > 1:
            required_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values
            required_values = [v.strip() for v in required_part.split(',')]
            # Check if ResMed value equals any required value
            for required in required_values:
                if required and resmed_val_str == required:
                    return True  # Equals required value
            return False  # Doesn't equal any required value
    
    return False


def check_if_condition_applies(rate_card_value, condition_text):
    """
    Check if a condition text contains a condition that applies to the given rate card value.
    This is used to determine if a failed condition match should disqualify a row.
    
    Args:
        rate_card_value: The value from Rate Card
        condition_text: The full condition rule text (may contain multiple conditions)
    
    Returns:
        True if the condition text contains a condition for this rate card value, False otherwise
    """
    if not condition_text or pd.isna(condition_text) or not rate_card_value or pd.isna(rate_card_value):
        return False
    
    rate_card_val_lower = str(rate_card_value).lower().strip()
    rate_card_val_normalized = rate_card_val_lower.replace(' ', '').replace('_', '')
    
    # Handle both string and list formats for conditions
    if isinstance(condition_text, str):
        conditions_list = [line.strip() for line in condition_text.split('\n') if line.strip()]
    elif isinstance(condition_text, list):
        conditions_list = condition_text
    else:
        conditions_list = [str(condition_text)]
    
    for cond_line in conditions_list:
        cond_lower = str(cond_line).lower()
        cond_normalized = cond_lower.replace(' ', '')
        
        # Skip header lines
        if 'conditional rules' in cond_lower and ':' not in cond_line:
            continue
        
        # Check if this condition line is for the rate card value
        # Pattern: "1. <value>:" or "<value>:"
        pattern_original = rf'(?:\d+\.\s*)?{re.escape(rate_card_val_lower)}:'
        pattern_normalized = rf'(?:\d+\.)?{re.escape(rate_card_val_normalized)}:'
        
        if re.search(pattern_original, cond_lower) or re.search(pattern_normalized, cond_normalized):
            return True  # Found a condition for this rate card value
    
    return False


def check_value_against_conditions(resmed_value, rate_card_value, column_name, conditions_dict):
    """Check if ResMed value satisfies any condition for the rate card value.
    
    Returns:
        (is_valid, matching_condition) tuple
    """
    # Try to find column in conditions_dict (case-insensitive)
    column_key = None
    for key in conditions_dict.keys():
        if normalize_column_name(key) == normalize_column_name(column_name):
            column_key = key
            break
    
    if column_key is None:
        return False, None
    
    conditions = conditions_dict[column_key]
    
    # Keep original rate card value (with spaces) for pattern matching
    rate_card_val_original = str(rate_card_value).strip() if pd.notna(rate_card_value) else ''
    rate_card_val_lower = rate_card_val_original.lower()
    
    # Also create a normalized version (no spaces) for fallback matching
    rate_card_val_normalized = rate_card_val_lower.replace(' ', '').replace('_', '')
    
    # Handle both string and list formats for conditions
    # rate_card_processing.py returns conditions as a string (from cell comments)
    if isinstance(conditions, str):
        # Split string by newlines to get individual condition lines
        conditions_list = [line.strip() for line in conditions.split('\n') if line.strip()]
    elif isinstance(conditions, list):
        conditions_list = conditions
    else:
        # If it's neither string nor list, try to convert
        conditions_list = [str(conditions)]
    
    for condition_text in conditions_list:
        # Check if this condition applies to the rate card value
        # Format: "1. NAC: RATE_TYPE is empty..." or "NAC: RATE_TYPE is empty..."
        condition_lower = str(condition_text).lower()
        
        # Skip header lines
        if 'conditional rules' in condition_lower and ':' not in condition_text:
            continue
        
        # Check if condition is for this rate card value
        # Look for pattern like "NAC:" or "1. NAC:" or "1.NAC:"
        if rate_card_val_lower:
            # Try matching with original formatting (with spaces)
            pattern_original = rf'(?:\d+\.\s*)?{re.escape(rate_card_val_lower)}:'
            
            # Also try normalized version (no spaces) for matching
            condition_normalized = condition_lower.replace(' ', '')
            pattern_normalized = rf'(?:\d+\.)?{re.escape(rate_card_val_normalized)}:'
            
            # Check both patterns
            matched = re.search(pattern_original, condition_lower) or re.search(pattern_normalized, condition_normalized)
            
            if matched:
                # This condition applies to this rate card value
                is_valid = value_satisfies_condition(resmed_value, rate_card_value, condition_text)
                if is_valid:
                    return True, condition_text
    
    return False, None


def load_standardized_dataframes():
    """Load standardized dataframes from shipments.py output."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "Standardized_Data.xlsx")
    
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found.")
        print("Please run shipments.py first to generate standardized data.")
        return None, None
    
    try:
        df_resmed = pd.read_excel(excel_file, sheet_name='ResMed (Standardized)')
        df_rate_card = pd.read_excel(excel_file, sheet_name='Rate Card (Standardized)')
        
        print(f"Loaded ResMed (Standardized): {df_resmed.shape}")
        print(f"Loaded Rate Card (Standardized): {df_rate_card.shape}")
        
        return df_resmed, df_rate_card
    except Exception as e:
        print(f"Error loading standardized data: {e}")
        return None, None


def find_common_columns(df_resmed, df_rate_card):
    """Find common columns between the two dataframes."""
    resmed_cols = set(df_resmed.columns)
    rate_card_cols = set(df_rate_card.columns)
    common_cols = sorted(list(resmed_cols & rate_card_cols))
    
    print(f"\nFound {len(common_cols)} common columns for matching:")
    for col in common_cols:
        print(f"  - {col}")
    
    return common_cols


def analyze_discrepancy_patterns(all_discrepancies):
    """
    Analyze discrepancies to find common patterns.
    
    Args:
        all_discrepancies: List of discrepancy dictionaries, each with 'column', 'etofs_value', 'rate_card_value'
    
    Returns:
        tuple: (has_common_pattern, pattern_comment)
            - has_common_pattern: True if there's a common pattern, False if all different
            - pattern_comment: The summarized comment based on the pattern
    """
    if not all_discrepancies:
        return False, "Please recheck the shipment details"
    
    # Group discrepancies by column name
    column_counts = {}
    column_discrepancies = {}
    
    for disc in all_discrepancies:
        col = disc.get('column', 'Unknown')
        if col not in column_counts:
            column_counts[col] = 0
            column_discrepancies[col] = []
        column_counts[col] += 1
        column_discrepancies[col].append(disc)
    
    total_discrepancies = len(all_discrepancies)
    unique_columns = len(column_counts)
    
    # If all discrepancies are for the same column - clear pattern
    if unique_columns == 1:
        column_name = list(column_counts.keys())[0]
        return True, f"{column_name}: Shipment value needs to be changed"
    
    # Check if one column dominates (has majority of discrepancies, at least 70%)
    for col, count in column_counts.items():
        if count / total_discrepancies >= 0.7:
            return True, f"{col}: Shipment value needs to be changed (and {total_discrepancies - count} other minor discrepancies)"
    
    # Check if a few columns (2-3) cover most discrepancies (80%+)
    sorted_columns = sorted(column_counts.items(), key=lambda x: x[1], reverse=True)
    top_columns = []
    covered_count = 0
    
    for col, count in sorted_columns[:3]:  # Check top 3 columns
        top_columns.append(col)
        covered_count += count
        if covered_count / total_discrepancies >= 0.8:
            break
    
    if len(top_columns) <= 3 and covered_count / total_discrepancies >= 0.8:
        # Format: "Column1, Column2: Shipment values need to be changed"
        columns_str = ", ".join(top_columns)
        return True, f"{columns_str}: Shipment values need to be changed"
    
    # No clear pattern - all different
    return False, "Please recheck the shipment details"


def match_shipments_with_rate_card(df_etofs, df_filtered_rate_card, common_columns, conditions_dict=None, business_rules_lookup=None):
    """Match shipments with Rate Card entries and find the corresponding Lane #.
    
    OPTIMIZED VERSION - Pre-parses conditions and uses fast lookups.
    Now includes business rules validation for exact lane matching.
    
    Args:
        df_etofs: Shipment dataframe (ETOF or LC) from vocabular.py
        df_filtered_rate_card: Rate Card standardized dataframe from rate_card_processing.py
        common_columns: List of common column names
        conditions_dict: Dictionary of conditional rules from rate_card_processing.py
                        Format: {column_name: condition_text, ...}
        business_rules_lookup: Dictionary with business rules for validating postal code zones
                              and country regions. Contains 'rule_to_country' and 'rule_to_postal_codes'
    """
    import time
    
    # Create a copy to preserve all original columns
    df_etofs = df_etofs.copy()
    
    total_shipments = len(df_etofs)
    total_rate_card_rows = len(df_filtered_rate_card)
    
    print(f"\n[DEBUG] Starting matching process:")
    print(f"[DEBUG]   - Total shipments to process: {total_shipments}")
    print(f"[DEBUG]   - Total rate card rows to compare: {total_rate_card_rows}")
    print(f"[DEBUG]   - Total comparisons: {total_shipments * total_rate_card_rows:,}")
    print(f"[DEBUG]   - Business rules available: {business_rules_lookup is not None}")
    
    # Create mappings from normalized column names back to original column names
    print(f"[DEBUG] Creating column mappings...")
    etofs_original_to_normalized = {col: normalize_column_name(col) for col in df_etofs.columns}
    rate_card_original_to_normalized = {col: normalize_column_name(col) for col in df_filtered_rate_card.columns}
    
    etofs_normalized_to_original = {v: k for k, v in etofs_original_to_normalized.items()}
    rate_card_normalized_to_original = {v: k for k, v in rate_card_original_to_normalized.items()}
    
    # Get normalized versions of common columns
    common_columns_normalized = [normalize_column_name(col) for col in common_columns]
    
    # Get the original column names for the common normalized columns
    common_etofs_cols_original = [etofs_normalized_to_original[col_norm] 
                                   for col_norm in common_columns_normalized 
                                   if col_norm in etofs_normalized_to_original]
    common_rate_card_cols_original = [rate_card_normalized_to_original[col_norm] 
                                      for col_norm in common_columns_normalized 
                                      if col_norm in rate_card_normalized_to_original]
    
    # PRE-PARSE all conditions into fast lookup structures
    print(f"[DEBUG] Pre-parsing conditions for fast lookup...")
    parsed_conditions = {}  # {col_norm: {rc_value_lower: {'type': ..., 'values': [...]}}}
    
    if conditions_dict:
        print(f"[DEBUG] CONDITIONS_DICT received with {len(conditions_dict)} entries:")
        for col_name, condition_text in conditions_dict.items():
            norm_col = normalize_column_name(col_name)
            parsed_conditions[norm_col] = parse_condition_text_fast(condition_text)
            print(f"   Column '{col_name}' (norm: '{norm_col}'):")
            print(f"      Raw condition: {condition_text[:100]}..." if len(condition_text) > 100 else f"      Raw condition: {condition_text}")
            print(f"      Parsed: {parsed_conditions[norm_col]}")
        print(f"[DEBUG] Parsed conditions for {len(parsed_conditions)} columns")
    else:
        print(f"[DEBUG] No conditions_dict provided!")
    
    # Identify business rule columns (columns containing postal code zones, country regions, etc.)
    business_rule_cols = set()
    business_rule_values = set()  # All business rule names (lowercase) for quick lookup
    if business_rules_lookup:
        br_cols = business_rules_lookup.get('business_rule_columns', set())
        for col in br_cols:
            business_rule_cols.add(normalize_column_name(col))
        # Get all rule names for quick lookup
        rule_to_country = business_rules_lookup.get('rule_to_country', {})
        rule_to_postal = business_rules_lookup.get('rule_to_postal_codes', {})
        for rule_name in rule_to_country.keys():
            business_rule_values.add(str(rule_name).lower().strip())
        for rule_name in rule_to_postal.keys():
            business_rule_values.add(str(rule_name).lower().strip())
        print(f"[DEBUG] Business rule columns: {business_rule_cols}")
        print(f"[DEBUG] Total business rule values: {len(business_rule_values)}")
        print(f"[DEBUG] Business rule values: {list(business_rule_values)[:20]}")
        
        # Show detailed business rules
        rule_to_excluded = business_rules_lookup.get('rule_to_excluded_postal_codes', {})
        print(f"\n[DEBUG] BUSINESS RULES LOADED:")
        for rule_name in list(rule_to_country.keys())[:15]:
            country = rule_to_country.get(rule_name, 'N/A')
            postal = rule_to_postal.get(rule_name, [])
            excluded = rule_to_excluded.get(rule_name, [])
            excluded_str = f", EXCLUDED={excluded[:5]}{'...' if len(excluded) > 5 else ''}" if excluded else ""
            print(f"   Rule '{rule_name}': Country={country}, PostalPrefixes={postal}{excluded_str}")
        if len(rule_to_country) > 15:
            print(f"   ... and {len(rule_to_country) - 15} more rules")
    else:
        print(f"[DEBUG] WARNING: business_rules_lookup is None - NO BUSINESS RULES AVAILABLE!")
    
    # DEBUG: Show all rate card columns vs common columns
    print(f"\n[DEBUG] ALL RATE CARD COLUMNS:")
    for col in df_filtered_rate_card.columns:
        in_common = col in common_columns or normalize_column_name(col) in common_columns_normalized
        marker = "✓" if in_common else "✗"
        print(f"   {marker} '{col}'")
    
    print(f"\n[DEBUG] COMMON COLUMNS BEING MATCHED:")
    for i, col in enumerate(common_columns):
        col_norm = common_columns_normalized[i] if i < len(common_columns_normalized) else "N/A"
        print(f"   {i+1}. '{col}' (normalized: '{col_norm}')")
    
    print(f"\nMatching based on {len(common_columns_normalized)} common attributes:")
    print(common_columns_normalized)
    
    # Identify business rule columns in rate card (even if not in common columns)
    # These are columns like "Origin Postal Code Zone", "Destination Postal Code Zone"
    rate_card_br_columns = []
    for col in df_filtered_rate_card.columns:
        col_lower = str(col).lower()
        if 'zone' in col_lower or 'region' in col_lower or 'hub' in col_lower:
            rate_card_br_columns.append(col)
    print(f"[DEBUG] Rate card business rule columns to check: {rate_card_br_columns}")
    
    # Identify Valid From and Valid To columns in rate card
    valid_from_col = None
    valid_to_col = None
    for col in df_filtered_rate_card.columns:
        col_lower = str(col).lower().replace(' ', '').replace('_', '')
        if 'validfrom' in col_lower:
            valid_from_col = col
        elif 'validto' in col_lower:
            valid_to_col = col
    
    if valid_from_col or valid_to_col:
        print(f"[DEBUG] Date validity columns found:")
        if valid_from_col:
            print(f"   - Valid From: '{valid_from_col}'")
        if valid_to_col:
            print(f"   - Valid To: '{valid_to_col}'")
    else:
        print(f"[DEBUG] No date validity columns found in rate card")
    
    # PRE-COMPUTE rate card normalized values ONCE
    print(f"\n[DEBUG] Pre-computing rate card values...")
    precompute_start = time.time()
    
    rate_card_precomputed = []
    for index_rate_card, row_rate_card in df_filtered_rate_card.iterrows():
        precomputed = {
            'index': index_rate_card,
            'row_dict': row_rate_card.to_dict(),
            'normalized_values': {},
            'raw_values': {},
            'raw_values_lower': {},  # Pre-compute lowercase for condition matching
            'br_column_values': {},  # Business rule column values (even if not in common columns)
            'valid_from': None,      # Valid From date
            'valid_to': None         # Valid To date
        }
        
        for i, col_norm in enumerate(common_columns_normalized):
            if i < len(common_rate_card_cols_original) and common_rate_card_cols_original[i] in row_rate_card:
                raw_val = row_rate_card[common_rate_card_cols_original[i]]
                precomputed['normalized_values'][col_norm] = normalize_value(raw_val)
                precomputed['raw_values'][col_norm] = raw_val
                # Pre-compute lowercase for condition matching
                if pd.notna(raw_val):
                    precomputed['raw_values_lower'][col_norm] = str(raw_val).lower().strip()
        
        # Also store business rule column values
        for br_col in rate_card_br_columns:
            if br_col in row_rate_card:
                precomputed['br_column_values'][br_col] = row_rate_card[br_col]
        
        # Store Valid From and Valid To dates
        if valid_from_col and valid_from_col in row_rate_card:
            precomputed['valid_from'] = row_rate_card[valid_from_col]
        if valid_to_col and valid_to_col in row_rate_card:
            precomputed['valid_to'] = row_rate_card[valid_to_col]
        
        rate_card_precomputed.append(precomputed)
    
    precompute_time = time.time() - precompute_start
    print(f"[DEBUG] Pre-computation completed in {precompute_time:.2f}s")
    
    # Pre-identify postal code columns
    postal_code_cols = set()
    for col_norm in common_columns_normalized:
        col_lower = col_norm.lower()
        if 'post' in col_lower or 'ship_post' in col_lower or 'cust_post' in col_lower:
            postal_code_cols.add(col_norm)
    
    # Identify SHIP_DATE column in shipment data
    ship_date_col = None
    ship_date_patterns = ['SHIP_DATE', 'ship_date', 'Ship Date', 'Loading date', 'SHIP DATE', 'ShipDate', 'shipdate']
    for col in df_etofs.columns:
        col_str = str(col)
        col_lower = col_str.lower().replace(' ', '').replace('_', '')
        for pattern in ship_date_patterns:
            pattern_lower = pattern.lower().replace(' ', '').replace('_', '')
            if col_lower == pattern_lower:
                ship_date_col = col
                break
        if ship_date_col:
            break
    
    if ship_date_col:
        print(f"[DEBUG] Ship date column found: '{ship_date_col}'")
    else:
        print(f"[DEBUG] No ship date column found in shipment data")
    
    # Start main matching loop
    print(f"\n[DEBUG] Starting main matching loop...")
    matching_start = time.time()
    last_progress_time = matching_start
    
    # Pre-allocate results list for faster assignment
    results = []
    
    # Convert to lists for faster iteration
    etofs_indices = df_etofs.index.tolist()
    etofs_rows = df_etofs.to_dict('records')
    # Keep original df rows for business rule validation (need Series with index)
    etofs_rows_series = [df_etofs.loc[idx] for idx in etofs_indices]
    
    # Enable detailed debug for first few shipments
    DEBUG_DETAILED = True
    DEBUG_MAX_SHIPMENTS = 12  # Show detailed debug for first N shipments
    
    for shipment_idx, (index_etofs, row_etofs_dict, row_etofs_series) in enumerate(zip(etofs_indices, etofs_rows, etofs_rows_series)):
        # Progress reporting every 20 shipments or every 10 seconds
        current_time = time.time()
        if shipment_idx % 20 == 0 or (current_time - last_progress_time) > 10:
            elapsed = current_time - matching_start
            rate = (shipment_idx + 1) / elapsed if elapsed > 0 else 0
            remaining = (total_shipments - shipment_idx - 1) / rate if rate > 0 else 0
            print(f"[DEBUG] Processing shipment {shipment_idx + 1}/{total_shipments} "
                  f"({100*(shipment_idx+1)/total_shipments:.1f}%) - "
                  f"Elapsed: {elapsed:.1f}s, Rate: {rate:.2f}/s, ETA: {remaining:.0f}s")
            last_progress_time = current_time
        
        # Detailed debug for first few shipments
        show_debug = DEBUG_DETAILED and shipment_idx < DEBUG_MAX_SHIPMENTS
        
        if show_debug:
            print(f"\n{'='*80}")
            print(f"[DETAILED DEBUG] SHIPMENT {shipment_idx + 1} (index {index_etofs})")
            print(f"{'='*80}")
            # Show shipment key values
            for key in ['Origin Country', 'Origin Postal Code', 'Destination Country', 'Destination Postal Code',
                       'SHIP_COUNTRY', 'SHIP_POST', 'CUST_COUNTRY', 'CUST_POST']:
                if key in row_etofs_dict:
                    print(f"   Shipment {key}: '{row_etofs_dict[key]}'")
        
        # Extract ship date from shipment
        ship_date_value = None
        if ship_date_col and ship_date_col in row_etofs_dict:
            ship_date_value = row_etofs_dict[ship_date_col]
            if show_debug and ship_date_value:
                print(f"   Shipment Ship Date: '{ship_date_value}'")
        
        # Prepare normalized values for the current ETOFS row
        etofs_normalized_values = {}
        etofs_raw_values_lower = {}
        
        for i, col_norm in enumerate(common_columns_normalized):
            if i < len(common_etofs_cols_original):
                col_orig = common_etofs_cols_original[i]
                if col_orig in row_etofs_dict:
                    raw_val = row_etofs_dict[col_orig]
                    etofs_normalized_values[col_norm] = normalize_value(raw_val)
                    if pd.notna(raw_val):
                        etofs_raw_values_lower[col_norm] = str(raw_val).lower().strip()
        
        max_matches = -1
        best_matching_rate_card_rows = []
        lanes_checked = []  # Track all lanes and their status for debug
        
        # Iterate through PRE-COMPUTED rate card rows
        for precomputed_rc in rate_card_precomputed:
            rate_card_normalized_values = precomputed_rc['normalized_values']
            rate_card_raw_values_lower = precomputed_rc['raw_values_lower']
            rate_card_raw_values = precomputed_rc['raw_values']
            
            lane_num = precomputed_rc['row_dict'].get('Lane #', precomputed_rc['row_dict'].get('Lane#', precomputed_rc['index']))
            
            current_matches = 0
            row_disqualified = False
            disqualify_reason = None
            match_details = []  # Track why each column matched/failed
            
            # CHECK DATE VALIDITY: Verify if shipment date falls within lane's validity period
            lane_valid_from = precomputed_rc.get('valid_from')
            lane_valid_to = precomputed_rc.get('valid_to')
            
            if lane_valid_from is not None or lane_valid_to is not None:
                date_is_valid, date_reason = is_date_in_validity_period(
                    ship_date_value, lane_valid_from, lane_valid_to, debug=show_debug
                )
                
                if date_is_valid is False:
                    # Date is outside validity period - disqualify this lane
                    row_disqualified = True
                    disqualify_reason = f"Date validity check failed: {date_reason}"
                    if show_debug:
                        match_details.append(f"   ✗ Lane {lane_num}: DATE VALIDITY FAILED → DISQUALIFIED")
                        match_details.append(f"      Valid From: {lane_valid_from}, Valid To: {lane_valid_to}")
                        match_details.append(f"      Ship Date: {ship_date_value}")
                        match_details.append(f"      Reason: {date_reason}")
                    
                    # Add to lanes_checked for debug
                    lanes_checked.append({
                        'lane': lane_num,
                        'matches': 0,
                        'disqualified': True,
                        'reason': disqualify_reason
                    })
                    continue  # Skip this lane
                elif date_is_valid is True and show_debug:
                    match_details.append(f"   ✓ Lane {lane_num}: DATE VALIDITY PASSED")
                    match_details.append(f"      {date_reason}")
            
            # Compare normalized values
            for i, col_norm in enumerate(common_columns_normalized):
                if row_disqualified:
                    break
                
                etofs_val = etofs_normalized_values.get(col_norm)
                rc_val = rate_card_normalized_values.get(col_norm)
                
                # Get original column name for business rule validation
                rate_card_original_col = common_rate_card_cols_original[i] if i < len(common_rate_card_cols_original) else None
                
                # WILDCARD: If rate card value is empty/None, any shipment value matches
                if rc_val is None:
                    current_matches += 1
                    if show_debug:
                        match_details.append(f"   ✓ {rate_card_original_col}: WILDCARD (empty) → +1")
                    continue
                
                if etofs_val is None and rc_val is None:
                    continue
                
                # CHECK BUSINESS RULES: If this column contains business rules (like Postal Code Zone)
                # and the rate card value is a business rule name, validate it
                if business_rules_lookup and rate_card_original_col:
                    rc_raw_val = rate_card_raw_values.get(col_norm)
                    rc_val_str = str(rc_raw_val).strip() if pd.notna(rc_raw_val) else ''
                    rc_val_lower = rc_val_str.lower()
                    
                    # Check if this rate card value is a business rule name
                    if rc_val_lower in business_rule_values:
                        # Validate business rule against shipment data
                        is_valid, reason = validate_business_rule_value(
                            rc_raw_val, row_etofs_series, business_rules_lookup, rate_card_original_col,
                            debug=show_debug
                        )
                        
                        if show_debug:
                            match_details.append(f"   [BR] {rate_card_original_col}='{rc_val_str}': is_valid={is_valid}, reason={reason}")
                        
                        if is_valid is True:
                            current_matches += 1
                            if show_debug:
                                match_details.append(f"   ✓ {rate_card_original_col}: BUSINESS RULE '{rc_val_str}' PASSED → +1")
                            continue
                        elif is_valid is False:
                            # Business rule validation failed - disqualify this lane
                            row_disqualified = True
                            disqualify_reason = f"Business rule '{rc_val_str}' failed: {reason}"
                            if show_debug:
                                match_details.append(f"   ✗ {rate_card_original_col}: BUSINESS RULE '{rc_val_str}' FAILED → DISQUALIFIED")
                                match_details.append(f"      Reason: {reason}")
                            continue
                        # If is_valid is None, fall through to other checks
                
                # Check conditions (fast path)
                if col_norm in parsed_conditions and col_norm in rate_card_raw_values_lower:
                    rc_val_lower = rate_card_raw_values_lower[col_norm]
                    etofs_val_lower = etofs_raw_values_lower.get(col_norm, '')
                    
                    if show_debug:
                        # Show detailed condition info
                        cond_info = parsed_conditions[col_norm]
                        match_details.append(f"   [COND CHECK] {rate_card_original_col}:")
                        match_details.append(f"      Rate card value: '{rc_val_lower}'")
                        match_details.append(f"      Shipment value: '{etofs_val_lower}'")
                        match_details.append(f"      Condition rules: {cond_info}")
                    
                    condition_result = check_condition_fast(
                        etofs_val_lower, rc_val_lower, parsed_conditions[col_norm], debug=show_debug
                    )
                    
                    if show_debug:
                        match_details.append(f"      Result: {condition_result}")
                    
                    if condition_result == 'matched':
                        current_matches += 1
                        if show_debug:
                            match_details.append(f"   ✓ {rate_card_original_col}: CONDITION matched → +1")
                        continue
                    elif condition_result == 'disqualified':
                        row_disqualified = True
                        disqualify_reason = f"Condition check failed for {rate_card_original_col}: rate_card='{rc_val_lower}', shipment='{etofs_val_lower}'"
                        if show_debug:
                            match_details.append(f"   ✗ {rate_card_original_col}: CONDITION DISQUALIFIED")
                        continue
                    # If 'no_condition', fall through to normal comparison
                
                # Normal comparison
                if col_norm in postal_code_cols and etofs_val and rc_val:
                    if str(etofs_val).startswith(str(rc_val)):
                        current_matches += 1
                        if show_debug:
                            match_details.append(f"   ✓ {rate_card_original_col}: POSTAL '{etofs_val}' starts with '{rc_val}' → +1")
                    else:
                        if show_debug:
                            match_details.append(f"   - {rate_card_original_col}: POSTAL '{etofs_val}' NOT starts with '{rc_val}' → +0")
                elif etofs_val == rc_val:
                    current_matches += 1
                    if show_debug:
                        match_details.append(f"   ✓ {rate_card_original_col}: EXACT '{etofs_val}' == '{rc_val}' → +1")
                else:
                    if show_debug:
                        match_details.append(f"   - {rate_card_original_col}: NO MATCH '{etofs_val}' != '{rc_val}' → +0")
            
            # ===== CHECK BUSINESS RULE COLUMNS (even if not in common columns) =====
            # These are columns like "Origin Postal Code Zone", "Destination Postal Code Zone"
            # that exist in rate card but not in shipment data
            if business_rules_lookup and not row_disqualified:
                br_column_values = precomputed_rc.get('br_column_values', {})
                
                for br_col, br_value in br_column_values.items():
                    if row_disqualified:
                        break
                    
                    # Skip if empty/None (wildcard)
                    if br_value is None or pd.isna(br_value) or str(br_value).strip() == '':
                        if show_debug:
                            match_details.append(f"   ✓ {br_col}: WILDCARD (empty) → +1")
                        current_matches += 1
                        continue
                    
                    br_value_str = str(br_value).strip()
                    br_value_lower = br_value_str.lower()
                    
                    # Check if this is a business rule
                    if br_value_lower in business_rule_values:
                        is_valid, reason = validate_business_rule_value(
                            br_value, row_etofs_series, business_rules_lookup, br_col,
                            debug=show_debug
                        )
                        
                        if show_debug:
                            match_details.append(f"   [BR CHECK] {br_col}='{br_value_str}': is_valid={is_valid}")
                        
                        if is_valid is True:
                            current_matches += 1
                            if show_debug:
                                match_details.append(f"   ✓ {br_col}: BUSINESS RULE '{br_value_str}' PASSED → +1")
                        elif is_valid is False:
                            # Business rule failed - DISQUALIFY this lane
                            row_disqualified = True
                            disqualify_reason = f"Business rule '{br_value_str}' failed: {reason}"
                            if show_debug:
                                match_details.append(f"   ✗ {br_col}: BUSINESS RULE '{br_value_str}' FAILED → DISQUALIFIED")
                                match_details.append(f"      Reason: {reason}")
                        else:
                            # is_valid is None - rule not found, treat as normal value
                            if show_debug:
                                match_details.append(f"   ? {br_col}: '{br_value_str}' not found as business rule → +0")
                    else:
                        # Not a business rule name - just log it
                        if show_debug:
                            match_details.append(f"   ? {br_col}: '{br_value_str}' not in business_rule_values → skipped")
            
            # Track this lane's status
            lane_status = {
                'lane': lane_num,
                'matches': current_matches,
                'disqualified': row_disqualified,
                'reason': disqualify_reason,
                'details': match_details
            }
            lanes_checked.append(lane_status)
            
            if row_disqualified:
                continue
            
            # Update best matches
            if current_matches > max_matches:
                max_matches = current_matches
                best_matching_rate_card_rows = [precomputed_rc['row_dict']]
            elif current_matches == max_matches and current_matches > 0:
                best_matching_rate_card_rows.append(precomputed_rc['row_dict'])
        
        # Show debug summary for this shipment
        if show_debug:
            print(f"\n[LANE SCORING SUMMARY] Shipment {shipment_idx + 1}:")
            print(f"   Total lanes checked: {len(lanes_checked)}")
            print(f"   Best match score: {max_matches}")
            print(f"   Best matching lanes: {len(best_matching_rate_card_rows)}")
            
            # Show disqualified lanes
            disqualified = [l for l in lanes_checked if l['disqualified']]
            if disqualified:
                print(f"\n   DISQUALIFIED LANES ({len(disqualified)}):")
                for l in disqualified[:10]:  # Show first 10
                    print(f"      Lane {l['lane']}: {l['reason']}")
            
            # Show top scoring lanes
            qualified = sorted([l for l in lanes_checked if not l['disqualified']], 
                             key=lambda x: x['matches'], reverse=True)
            if qualified:
                print(f"\n   TOP SCORING LANES:")
                for l in qualified[:10]:  # Show top 10
                    marker = "★" if l['matches'] == max_matches else " "
                    print(f"      {marker} Lane {l['lane']}: {l['matches']} matches")
                    # Show details for best matching lanes
                    if l['matches'] == max_matches:
                        for detail in l['details']:
                            print(f"         {detail}")
            
            print(f"{'='*80}\n")
        
        # Build comment
        if len(best_matching_rate_card_rows) == 0:
            comment = "No matching rate card entries found"
        elif len(best_matching_rate_card_rows) == 1:
            lane_num = best_matching_rate_card_rows[0].get('Lane #', best_matching_rate_card_rows[0].get('Lane#', 'N/A'))
            comment = f"Rate lane: {lane_num}"
        else:
            lane_nums = []
            for rc_row in best_matching_rate_card_rows:
                lane_num = rc_row.get('Lane #', rc_row.get('Lane#', 'N/A'))
                if lane_num not in lane_nums:
                    lane_nums.append(str(lane_num))
            comment = f"Rate lanes: {', '.join(lane_nums)}"
        
        results.append((index_etofs, comment))
    
    # Batch assign all comments at once (faster than individual assignments)
    for index_etofs, comment in results:
        df_etofs.loc[index_etofs, 'comment'] = comment
    
    total_time = time.time() - matching_start
    print(f"\n[DEBUG] Matching completed:")
    print(f"[DEBUG]   - Total time: {total_time:.2f}s")
    print(f"[DEBUG]   - Shipments processed: {len(results)}")
    if len(results) > 0:
        print(f"[DEBUG]   - Average time per shipment: {total_time/len(results):.3f}s")
    
    return df_etofs


def parse_condition_text_fast(condition_text):
    """
    Pre-parse condition text into a fast lookup structure.
    Returns: {rc_value_lower: {'type': 'equals'|'not_equals'|'contains'|'not_contains'|'is_empty', 'values': [...]}}
    """
    if not condition_text or pd.isna(condition_text):
        return {}
    
    result = {}
    condition_text = str(condition_text)
    
    # Split by newlines to get individual condition lines
    lines = [line.strip() for line in condition_text.split('\n') if line.strip()]
    
    for line in lines:
        line_lower = line.lower()
        
        # Skip header lines
        if 'conditional rules' in line_lower and ':' not in line:
            continue
        
        # Parse format: "1. VALUE: condition" or "VALUE: condition"
        if ':' not in line:
            continue
        
        # Remove leading number (e.g., "1. " or "2.")
        cleaned_line = re.sub(r'^\d+\.\s*', '', line)
        parts = cleaned_line.split(':', 1)
        
        if len(parts) < 2:
            continue
        
        rc_value = parts[0].strip().lower()
        condition_logic = parts[1].strip().lower()
        
        if not rc_value:
            continue
        
        # Determine condition type and extract values
        if 'is empty' in condition_logic:
            result[rc_value] = {'type': 'is_empty', 'values': []}
        elif 'does not equal' in condition_logic or 'does not equal to' in condition_logic:
            # Extract values after "does not equal" or "does not equal to"
            match = re.search(r'does not equal(?:\s+to)?\s+(.+?)(?:\s+in\s+any\s+item)?$', condition_logic)
            if match:
                values = [v.strip().lower() for v in match.group(1).split(',')]
                result[rc_value] = {'type': 'not_equals', 'values': values}
        elif 'does not contain' in condition_logic:
            match = re.search(r'does not contain\s+(.+?)(?:\s+in\s+any\s+item)?$', condition_logic)
            if match:
                values = [v.strip().lower() for v in match.group(1).split(',')]
                result[rc_value] = {'type': 'not_contains', 'values': values}
        elif 'equal to' in condition_logic or 'equals' in condition_logic:
            if 'equal to' in condition_logic:
                match = re.search(r'equal(?:s)?\s+to\s+(.+?)(?:\s+in\s+any\s+item)?$', condition_logic)
            else:
                match = re.search(r'equals\s+(.+?)(?:\s+in\s+any\s+item)?$', condition_logic)
            if match:
                values = [v.strip().lower() for v in match.group(1).split(',')]
                result[rc_value] = {'type': 'equals', 'values': values}
        elif 'contains' in condition_logic:
            match = re.search(r'contains\s+(.+?)(?:\s+in\s+any\s+item)?$', condition_logic)
            if match:
                values = [v.strip().lower() for v in match.group(1).split(',')]
                result[rc_value] = {'type': 'contains', 'values': values}
    
    return result


def check_condition_fast(shipment_val_lower, rc_val_lower, parsed_conditions, debug=False):
    """
    Fast condition checking using pre-parsed conditions.
    
    Returns:
        'matched' - condition found and satisfied
        'disqualified' - condition found but NOT satisfied (reject this rate card row)
        'no_condition' - no condition applies to this rate card value
    """
    # Check if there's a condition for this rate card value
    if rc_val_lower not in parsed_conditions:
        if debug:
            print(f"         [check_condition_fast] rc_val='{rc_val_lower}' NOT in parsed_conditions keys: {list(parsed_conditions.keys())[:5]}")
        return 'no_condition'
    
    cond = parsed_conditions[rc_val_lower]
    cond_type = cond['type']
    cond_values = cond['values']
    
    if debug:
        print(f"         [check_condition_fast] Found condition for '{rc_val_lower}': type={cond_type}, values={cond_values}")
    
    # Check if shipment value is empty
    is_empty = not shipment_val_lower or shipment_val_lower in ('', 'nan', 'none', 'null')
    
    if debug:
        print(f"         [check_condition_fast] shipment_val='{shipment_val_lower}', is_empty={is_empty}")
    
    if cond_type == 'is_empty':
        result = 'matched' if is_empty else 'disqualified'
        if debug:
            print(f"         [check_condition_fast] is_empty check: {result}")
        return result
    
    if cond_type == 'not_equals':
        if is_empty:
            return 'matched'  # Empty doesn't equal anything
        for forbidden in cond_values:
            if shipment_val_lower == forbidden:
                if debug:
                    print(f"         [check_condition_fast] not_equals: shipment equals forbidden '{forbidden}' → disqualified")
                return 'disqualified'
        return 'matched'
    
    if cond_type == 'not_contains':
        if is_empty:
            return 'matched'  # Empty doesn't contain anything
        for forbidden in cond_values:
            if forbidden in shipment_val_lower:
                if debug:
                    print(f"         [check_condition_fast] not_contains: shipment contains forbidden '{forbidden}' → disqualified")
                return 'disqualified'
        return 'matched'
    
    if cond_type == 'equals':
        if is_empty:
            return 'disqualified'
        for required in cond_values:
            if shipment_val_lower == required:
                return 'matched'
        return 'disqualified'
    
    if cond_type == 'contains':
        if is_empty:
            return 'disqualified'
        for required in cond_values:
            if required in shipment_val_lower:
                return 'matched'
        return 'disqualified'
    
    return 'no_condition'


def get_partly_df_folder():
    """Get the path to the partly_df folder."""
    return Path(__file__).parent / "partly_df"


def discover_vocabulary_mapping_files():
    """
    Discover all vocabulary mapping files in partly_df/ folder.
    These are files created by vocabular.py with pattern: <agreement>_vocabulary_mapping.xlsx
    
    Returns:
        dict: {agreement_number: file_path, ...}
    """
    partly_df = get_partly_df_folder()
    if not partly_df.exists():
        print(f"   [ERROR] partly_df folder not found: {partly_df}")
        return {}
    
    mapping_files = {}
    for file in partly_df.glob("*_vocabulary_mapping.xlsx"):
        # Extract agreement number from filename (e.g., "RA20241129009_vocabulary_mapping.xlsx" -> "RA20241129009")
        agreement_number = file.stem.replace("_vocabulary_mapping", "")
        mapping_files[agreement_number] = file
    
    return mapping_files


def discover_rate_card_files():
    """
    Discover all rate card files in partly_df/ folder.
    These are files created by part4_rate_card_processing.py with pattern: <agreement>.xlsx
    
    Returns:
        dict: {agreement_number: file_path, ...}
    """
    partly_df = get_partly_df_folder()
    if not partly_df.exists():
        print(f"   [ERROR] partly_df folder not found: {partly_df}")
        return {}
    
    rate_card_files = {}
    for file in partly_df.glob("*.xlsx"):
        # Skip vocabulary mapping files and other non-rate-card files
        if "_vocabulary_mapping" in file.stem:
            continue
        if "_matched" in file.stem:
            continue
        if "lc_etof_mapping" in file.stem.lower():
            continue
        if "order_lc_etof_mapping" in file.stem.lower():
            continue
        
        # Check if it's a rate card file (has "Rate Card Data" sheet)
        try:
            xl = pd.ExcelFile(file)
            if "Rate Card Data" in xl.sheet_names:
                agreement_number = file.stem
                rate_card_files[agreement_number] = file
        except Exception:
            pass
    
    return rate_card_files


def read_rate_card_from_partly_df(rate_card_file_path):
    """
    Read rate card data from a file in partly_df/ folder.
    
    Args:
        rate_card_file_path: Path to the rate card file
    
    Returns:
        tuple: (dataframe, columns, conditions, business_rules_lookup) or (None, [], {}, None) if error
    """
    try:
        df = pd.read_excel(rate_card_file_path, sheet_name="Rate Card Data")
        columns = df.columns.tolist()
        
        # Try to read conditions
        conditions = {}
        try:
            df_conditions = pd.read_excel(rate_card_file_path, sheet_name="Conditions")
            for _, row in df_conditions.iterrows():
                col_name = row.get("Column")
                condition = row.get("Condition Rule", "")
                if col_name and condition and str(condition).strip():
                    conditions[col_name] = str(condition).strip()
        except Exception:
            pass
        
        # Try to read business rules
        business_rules_lookup = None
        try:
            xl = pd.ExcelFile(rate_card_file_path)
            if "Business Rules" in xl.sheet_names:
                df_br = pd.read_excel(rate_card_file_path, sheet_name="Business Rules")
                business_rules_lookup = parse_business_rules_from_sheet(df_br, df, columns)
                if business_rules_lookup:
                    print(f"   [Business Rules] Loaded from rate card file")
        except Exception as e:
            print(f"   [DEBUG] Could not read business rules from rate card: {e}")
        
        return df, columns, conditions, business_rules_lookup
    except Exception as e:
        print(f"   [ERROR] Could not read rate card: {e}")
        return None, [], {}, None


def parse_business_rules_from_sheet(df_business_rules, rate_card_df, rate_card_columns):
    """
    Parse business rules from a Business Rules sheet and create lookup structures.
    
    Args:
        df_business_rules: DataFrame from Business Rules sheet
        rate_card_df: DataFrame with rate card data
        rate_card_columns: List of rate card column names
    
    Returns:
        dict: Business rules lookup with rule_to_country, rule_to_postal_codes, etc.
    """
    if df_business_rules is None or df_business_rules.empty:
        return None
    
    result = {
        'rule_to_country': {},
        'rule_to_postal_codes': {},
        'business_rule_columns': set(),
        'column_to_rules': {},
        'all_rules': []
    }
    
    # Parse the business rules sheet
    # Typical format: Rule Name, Country, Postal Codes (comma-separated or multiple rows)
    current_rule = None
    current_country = None
    current_postal_codes = []
    
    for _, row in df_business_rules.iterrows():
        # Try to find rule name column
        rule_name = None
        country = None
        postal_codes = []
        
        for col in df_business_rules.columns:
            col_lower = str(col).lower()
            val = row.get(col)
            
            if pd.isna(val) or str(val).strip() == '':
                continue
            
            val_str = str(val).strip()
            
            # Identify column type
            if 'name' in col_lower or 'rule' in col_lower or 'zone' in col_lower or 'region' in col_lower:
                rule_name = val_str
            elif 'country' in col_lower:
                country = val_str
            elif 'postal' in col_lower or 'code' in col_lower or 'prefix' in col_lower:
                # Handle comma-separated postal codes
                if ',' in val_str:
                    postal_codes.extend([p.strip() for p in val_str.split(',') if p.strip()])
                else:
                    postal_codes.append(val_str)
        
        # If we found a rule name, save it
        if rule_name:
            if current_rule and current_rule != rule_name:
                # Save previous rule
                if current_country:
                    result['rule_to_country'][current_rule] = current_country
                if current_postal_codes:
                    result['rule_to_postal_codes'][current_rule] = current_postal_codes
                result['all_rules'].append({
                    'name': current_rule,
                    'country': current_country,
                    'postal_codes': current_postal_codes
                })
            
            current_rule = rule_name
            current_country = country if country else current_country
            current_postal_codes = postal_codes if postal_codes else []
        elif current_rule:
            # Continuation of previous rule - add postal codes
            if country:
                current_country = country
            if postal_codes:
                current_postal_codes.extend(postal_codes)
    
    # Save last rule
    if current_rule:
        if current_country:
            result['rule_to_country'][current_rule] = current_country
        if current_postal_codes:
            result['rule_to_postal_codes'][current_rule] = current_postal_codes
        result['all_rules'].append({
            'name': current_rule,
            'country': current_country,
            'postal_codes': current_postal_codes
        })
    
    # Find which columns contain business rule values
    rule_names_lower = {str(r).lower() for r in result['rule_to_country'].keys()}
    rule_names_lower.update({str(r).lower() for r in result['rule_to_postal_codes'].keys()})
    
    for col in rate_card_columns:
        col_lower = str(col).lower()
        # Check if column name suggests business rules
        if 'zone' in col_lower or 'region' in col_lower or 'hub' in col_lower:
            result['business_rule_columns'].add(col)
            # Find all unique values in this column that are rule names
            if col in rate_card_df.columns:
                unique_vals = rate_card_df[col].dropna().unique()
                rules_in_col = []
                for val in unique_vals:
                    val_lower = str(val).lower().strip()
                    if val_lower in rule_names_lower:
                        rules_in_col.append(str(val).strip())
                if rules_in_col:
                    result['column_to_rules'][col] = rules_in_col
    
    if result['rule_to_country'] or result['rule_to_postal_codes']:
        print(f"   [Business Rules] Parsed {len(result['rule_to_country'])} rules with country, {len(result['rule_to_postal_codes'])} with postal codes")
        print(f"   [Business Rules] Columns with rules: {result['business_rule_columns']}")
        return result
    
    return None


def read_vocabulary_mapping_from_partly_df(mapping_file_path):
    """
    Read vocabulary mapping data from a file in partly_df/ folder.
    
    Args:
        mapping_file_path: Path to the vocabulary mapping file
    
    Returns:
        tuple: (dataframe, columns) or (None, []) if error
    """
    try:
        # Try to read "Mapped Data" sheet first, then "LC" sheet
        xl = pd.ExcelFile(mapping_file_path)
        
        if "Mapped Data" in xl.sheet_names:
            df = pd.read_excel(mapping_file_path, sheet_name="Mapped Data")
        elif "LC" in xl.sheet_names:
            df = pd.read_excel(mapping_file_path, sheet_name="LC")
        else:
            # Use first sheet
            df = pd.read_excel(mapping_file_path, sheet_name=0)
        
        columns = df.columns.tolist()
        return df, columns
    except Exception as e:
        print(f"   [ERROR] Could not read vocabulary mapping: {e}")
        return None, []


def get_lc_etof_mapping_file():
    """Find the LC/ETOF mapping file in partly_df folder."""
    partly_df = get_partly_df_folder()
    
    # Try different possible names
    possible_names = ["lc_etof_mapping.xlsx", "order_lc_etof_mapping.xlsx"]
    for name in possible_names:
        file_path = partly_df / name
        if file_path.exists():
            return file_path
    
    return None


def read_lc_etof_mapping_for_agreement(agreement_number):
    """
    Read LC/ETOF mapping data for a specific agreement from the mapping file.
    
    Args:
        agreement_number: The agreement number (tab name in the mapping file)
    
    Returns:
        tuple: (dataframe, columns) or (None, []) if not found
    """
    mapping_file = get_lc_etof_mapping_file()
    if mapping_file is None:
        print(f"   [WARNING] LC/ETOF mapping file not found in partly_df/")
        return None, []
    
    try:
        xl = pd.ExcelFile(mapping_file)
        
        # Check if this agreement has a tab
        if agreement_number in xl.sheet_names:
            df = pd.read_excel(mapping_file, sheet_name=agreement_number)
            return df, df.columns.tolist()
        else:
            print(f"   [WARNING] Tab '{agreement_number}' not found in LC/ETOF mapping file")
            return None, []
    except Exception as e:
        print(f"   [ERROR] Could not read LC/ETOF mapping: {e}")
        return None, []


def update_lc_etof_mapping_with_comments(agreement_number, df_matched):
    """
    Update the lc_etof_mapping.xlsx file by adding the 'comment' column from matched data
    to the corresponding agreement tab, matching on ETOF #.
    
    Args:
        agreement_number: The agreement number (tab name in lc_etof_mapping.xlsx)
        df_matched: The matched DataFrame with 'comment' column
    
    Returns:
        bool: True if successful, False otherwise
    """
    print(f"\n   Updating lc_etof_mapping.xlsx with comments for {agreement_number}...")
    
    if df_matched is None or df_matched.empty:
        print(f"   [WARNING] No matched data to update")
        return False
    
    # Check if 'comment' column exists
    if 'comment' not in df_matched.columns:
        print(f"   [WARNING] No 'comment' column in matched data")
        return False
    
    # Find ETOF # column in matched data
    etof_col_matched = None
    etof_variations = ['ETOF #', 'ETOF#', 'etof #', 'etof#']
    for col in df_matched.columns:
        for var in etof_variations:
            if col.lower().replace(' ', '') == var.lower().replace(' ', ''):
                etof_col_matched = col
                break
        if etof_col_matched:
            break
    
    if not etof_col_matched:
        print(f"   [WARNING] ETOF # column not found in matched data")
        return False
    
    # Get the LC/ETOF mapping file
    mapping_file = get_lc_etof_mapping_file()
    if mapping_file is None:
        print(f"   [WARNING] lc_etof_mapping.xlsx not found")
        return False
    
    try:
        # Read all sheets from the mapping file
        xl = pd.ExcelFile(mapping_file)
        all_sheets = {}
        
        for sheet_name in xl.sheet_names:
            all_sheets[sheet_name] = pd.read_excel(mapping_file, sheet_name=sheet_name)
        
        # Check if agreement tab exists
        if agreement_number not in all_sheets:
            print(f"   [WARNING] Tab '{agreement_number}' not found in lc_etof_mapping.xlsx")
            return False
        
        df_tab = all_sheets[agreement_number]
        
        # Find ETOF # column in tab
        etof_col_tab = None
        for col in df_tab.columns:
            for var in etof_variations:
                if col.lower().replace(' ', '') == var.lower().replace(' ', ''):
                    etof_col_tab = col
                    break
            if etof_col_tab:
                break
        
        if not etof_col_tab:
            print(f"   [WARNING] ETOF # column not found in tab '{agreement_number}'")
            return False
        
        # Create mapping from ETOF # to comment
        etof_to_comment = {}
        for _, row in df_matched.iterrows():
            etof_val = row.get(etof_col_matched)
            comment_val = row.get('comment')
            if pd.notna(etof_val) and str(etof_val).strip() and str(etof_val).lower() != 'nan':
                etof_key = str(etof_val).strip()
                if pd.notna(comment_val):
                    etof_to_comment[etof_key] = str(comment_val)
        
        print(f"   Found {len(etof_to_comment)} comments to add")
        
        # Add/update comment column in the tab
        if 'comment' not in df_tab.columns:
            df_tab['comment'] = None
        
        # Update comments based on ETOF #
        updated_count = 0
        for idx, row in df_tab.iterrows():
            etof_val = row.get(etof_col_tab)
            if pd.notna(etof_val) and str(etof_val).strip() and str(etof_val).lower() != 'nan':
                etof_key = str(etof_val).strip()
                if etof_key in etof_to_comment:
                    df_tab.at[idx, 'comment'] = etof_to_comment[etof_key]
                    updated_count += 1
        
        print(f"   Updated {updated_count} rows with comments")
        
        # Update the sheet in all_sheets
        all_sheets[agreement_number] = df_tab
        
        # Also update 'All Data' tab if it exists
        if 'All Data' in all_sheets:
            df_all = all_sheets['All Data']
            
            # Find ETOF # column in All Data
            etof_col_all = None
            for col in df_all.columns:
                for var in etof_variations:
                    if col.lower().replace(' ', '') == var.lower().replace(' ', ''):
                        etof_col_all = col
                        break
                if etof_col_all:
                    break
            
            if etof_col_all:
                if 'comment' not in df_all.columns:
                    df_all['comment'] = None
                
                all_data_updated = 0
                for idx, row in df_all.iterrows():
                    etof_val = row.get(etof_col_all)
                    if pd.notna(etof_val) and str(etof_val).strip() and str(etof_val).lower() != 'nan':
                        etof_key = str(etof_val).strip()
                        if etof_key in etof_to_comment:
                            df_all.at[idx, 'comment'] = etof_to_comment[etof_key]
                            all_data_updated += 1
                
                all_sheets['All Data'] = df_all
                print(f"   Also updated {all_data_updated} rows in 'All Data' tab")
        
        # Save back to the file
        with pd.ExcelWriter(mapping_file, engine='openpyxl') as writer:
            for sheet_name, df_sheet in all_sheets.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"   ✓ Updated lc_etof_mapping.xlsx successfully")
        return True
        
    except Exception as e:
        print(f"   [ERROR] Failed to update lc_etof_mapping.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return False


def run_matching_for_agreement(agreement_number, rate_card_file, mapping_file):
    """
    Run matching for a single agreement.
    
    Args:
        agreement_number: The agreement number (e.g., "RA20241129009")
        rate_card_file: Path to the rate card file
        mapping_file: Path to the vocabulary mapping file
    
    Returns:
        str: Path to output file if successful, None otherwise
    """
    print(f"\n{'='*80}")
    print(f"PROCESSING AGREEMENT: {agreement_number}")
    print(f"{'='*80}")
    
    # Step 1: Load rate card
    print(f"\n1. Loading rate card from: {rate_card_file.name}")
    df_rate_card, rate_card_columns, rate_card_conditions, business_rules_lookup = read_rate_card_from_partly_df(rate_card_file)
    
    if df_rate_card is None or df_rate_card.empty:
        print(f"   [ERROR] Could not load rate card for {agreement_number}")
        return None
    
    print(f"   Loaded: {len(df_rate_card)} rows, {len(rate_card_columns)} columns")
    print(f"   Conditions loaded: {len(rate_card_conditions)} columns with conditional rules")
    if rate_card_conditions:
        for col_name in list(rate_card_conditions.keys())[:5]:  # Show first 5
            cond_text = rate_card_conditions[col_name]
            print(f"     - {col_name}: {cond_text[:60]}..." if len(cond_text) > 60 else f"     - {col_name}: {cond_text}")
        if len(rate_card_conditions) > 5:
            print(f"     ... and {len(rate_card_conditions) - 5} more")
    
    # Step 1b: Try to load business rules from original rate card if not in partly_df file
    if business_rules_lookup is None and BUSINESS_RULES_AVAILABLE:
        # Try to find original rate card in input folder
        input_folder = Path(__file__).parent / "input"
        if input_folder.exists():
            # Look for rate card files with matching agreement number
            for file in input_folder.glob("*.xlsx"):
                if agreement_number in file.stem or 'rate' in file.stem.lower():
                    try:
                        print(f"\n1b. Trying to load business rules from: {file.name}")
                        business_rules_lookup = load_business_rules_for_matching(file.name)
                        if business_rules_lookup:
                            break
                    except Exception as e:
                        print(f"   [DEBUG] Could not load business rules from {file.name}: {e}")
    
    # Step 2: Load vocabulary mapping (shipment data)
    print(f"\n2. Loading vocabulary mapping from: {mapping_file.name}")
    df_shipments, shipment_columns = read_vocabulary_mapping_from_partly_df(mapping_file)
    
    if df_shipments is None or df_shipments.empty:
        print(f"   [ERROR] Could not load vocabulary mapping for {agreement_number}")
        return None
    
    print(f"   Loaded: {len(df_shipments)} rows, {len(shipment_columns)} columns")
    
    # Step 3: Filter rows with ETOF # values
    print(f"\n3. Filtering rows with ETOF # values...")
    etof_col = None
    etof_col_variations = ['ETOF #', 'ETOF#', 'etof #', 'etof#', 'ETOF', 'etof']
    
    for col in df_shipments.columns:
        col_normalized = str(col).strip()
        for variation in etof_col_variations:
            if col_normalized.lower() == variation.lower() or col_normalized.lower().replace(' ', '') == variation.lower().replace(' ', ''):
                etof_col = col
                break
        if etof_col:
            break
    
    if etof_col:
        initial_count = len(df_shipments)
        df_shipments = df_shipments[df_shipments[etof_col].notna()]
        df_shipments = df_shipments[df_shipments[etof_col].astype(str).str.strip() != '']
        df_shipments = df_shipments[df_shipments[etof_col].astype(str).str.lower() != 'nan']
        print(f"   Filtered: {initial_count} -> {len(df_shipments)} rows")
    
    if df_shipments.empty:
        print(f"   [WARNING] No rows with ETOF # values for {agreement_number}")
        return None
    
    # Step 4: Find common columns
    print(f"\n4. Finding common columns...")
    common_columns = find_common_columns(df_shipments, df_rate_card)
    
    if not common_columns:
        print(f"   [WARNING] No common columns found for {agreement_number}")
        return None
    
    # Step 5: Match shipments with rate card
    print(f"\n5. Matching shipments with rate card...")
    df_result = match_shipments_with_rate_card(df_shipments, df_rate_card, common_columns, rate_card_conditions, business_rules_lookup)
    
    # Step 6: Save results
    print(f"\n6. Saving results...")
    output_file = get_partly_df_folder() / f"{agreement_number}_matched.xlsx"
    
    try:
        # Reorder columns
        df_result_reordered = reorder_columns_for_output(df_result)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_result_reordered.to_excel(writer, sheet_name='Matched Shipments', index=False)
            
            # Apply formatting
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                workbook = writer.book
                ws = workbook['Matched Shipments']
                
                # Style header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                ws.freeze_panes = 'A2'
            except Exception:
                pass
        
        print(f"   Saved to: {output_file}")
        
        # Show summary
        rows_with_lane = df_result[df_result['comment'].str.startswith('Rate lane', na=False)]
        rows_no_match = df_result[df_result['comment'] == 'No matching rate card entries found']
        print(f"   - Rows with Rate Lane: {len(rows_with_lane)}")
        print(f"   - Rows with no match: {len(rows_no_match)}")
        
        return str(output_file)
        
    except Exception as e:
        print(f"   [ERROR] Failed to save: {e}")
        return None


def reorder_columns_for_output(df):
    """Reorder columns with priority: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date, then others."""
    if df is None or df.empty:
        return df
    
    # Find priority columns
    lc_patterns = ['LC #', 'LC#', 'lc #', 'lc#']
    etof_patterns = ['ETOF #', 'ETOF#', 'etof #', 'etof#']
    shipment_id_patterns = ['Shipment ID', 'ShipmentID', 'SHIPMENT_ID']
    delivery_patterns = ['Delivery Number', 'DeliveryNumber', 'DELIVERY_NUMBER', 'DELIVERY NUMBER(s)']
    carrier_patterns = ['Carrier', 'carrier', 'CARRIER']
    ship_date_patterns = ['SHIP_DATE', 'ship_date', 'Ship Date', 'Loading date']
    carrier_agreement_patterns = ['Carrier agreement #', 'Carrier Agreement #']
    
    lc_col = etof_col = shipment_id_col = delivery_col = carrier_col = ship_date_col = carrier_agreement_col = None
    
    for col in df.columns:
        col_str = str(col).strip()
        col_lower = col_str.lower().replace(' ', '').replace('_', '')
        
        if not lc_col:
            for pattern in lc_patterns:
                if col_lower == pattern.lower().replace(' ', '') or col_str == pattern:
                    lc_col = col
                    break
        
        if not etof_col:
            for pattern in etof_patterns:
                if col_lower == pattern.lower().replace(' ', '') or col_str == pattern:
                    etof_col = col
                    break
        
        if not shipment_id_col:
            for pattern in shipment_id_patterns:
                if col_lower == pattern.lower().replace(' ', '').replace('_', '') or col_str == pattern:
                    shipment_id_col = col
                    break
        
        if not delivery_col:
            for pattern in delivery_patterns:
                if col_lower == pattern.lower().replace(' ', '').replace('_', '') or 'delivery' in col_lower and 'number' in col_lower:
                    delivery_col = col
                    break
        
        if not carrier_col:
            for pattern in carrier_patterns:
                if col_lower == pattern.lower() or col_str == pattern:
                    carrier_col = col
                    break
        
        if not ship_date_col:
            for pattern in ship_date_patterns:
                if col_lower == pattern.lower().replace(' ', '_') or col_str == pattern:
                    ship_date_col = col
                    break
        
        if not carrier_agreement_col:
            for pattern in carrier_agreement_patterns:
                if pattern.lower().replace(' ', '') in col_lower:
                    carrier_agreement_col = col
                    break
    
    # Build ordered column list
    ordered_columns = []
    priority_cols = [lc_col, etof_col, shipment_id_col, delivery_col, carrier_col, carrier_agreement_col, ship_date_col]
    
    for col in priority_cols:
        if col:
            ordered_columns.append(col)
    
    priority_set = set(ordered_columns)
    for col in df.columns:
        if col not in priority_set and col != 'comment':
            ordered_columns.append(col)
    
    if 'comment' in df.columns:
        ordered_columns.append('comment')
    
    return df[ordered_columns]


def run_matching_all_agreements():
    """
    Run matching for all agreements found in partly_df/ folder.
    
    Workflow:
    1. Discover vocabulary mapping files (created by vocabular.py)
    2. Discover rate card files (created by part4_rate_card_processing.py)
    3. For each agreement that has both files, run matching
    4. Create <agreement>_matched.xlsx files
    5. Update lc_etof_mapping.xlsx with comments from matched data
    
    Returns:
        dict: {agreement_number: matched_file_path, ...}
    """
    print("\n" + "="*80)
    print("RATE LANE MATCHING - MULTI AGREEMENT PROCESSING")
    print("="*80)
    
    # Step 1: Discover files
    print("\n1. Discovering files in partly_df/ folder...")
    
    mapping_files = discover_vocabulary_mapping_files()
    rate_card_files = discover_rate_card_files()
    
    print(f"\n   Vocabulary mapping files found ({len(mapping_files)}):")
    for agreement, file in mapping_files.items():
        print(f"     - {agreement}: {file.name}")
    
    print(f"\n   Rate card files found ({len(rate_card_files)}):")
    for agreement, file in rate_card_files.items():
        print(f"     - {agreement}: {file.name}")
    
    # Step 2: Find agreements that have both files
    common_agreements = set(mapping_files.keys()) & set(rate_card_files.keys())
    
    if not common_agreements:
        print("\n   [ERROR] No agreements found with both vocabulary mapping and rate card files!")
        print("   Please ensure:")
        print("     1. part4_rate_card_processing.py has been run (creates <agreement>.xlsx)")
        print("     2. vocabular.py has been run (creates <agreement>_vocabulary_mapping.xlsx)")
        return {}
    
    print(f"\n   Agreements to process ({len(common_agreements)}):")
    for agreement in sorted(common_agreements):
        print(f"     - {agreement}")
    
    # Step 3: Process each agreement
    results = {}
    
    for i, agreement in enumerate(sorted(common_agreements), 1):
        print(f"\n[{i}/{len(common_agreements)}] Processing: {agreement}")
        
        matched_file = run_matching_for_agreement(
            agreement_number=agreement,
            rate_card_file=rate_card_files[agreement],
            mapping_file=mapping_files[agreement]
        )
        
        if matched_file:
            results[agreement] = matched_file
    
    # Summary
    print(f"\n{'='*80}")
    print(f"MATCHING COMPLETE")
    print(f"{'='*80}")
    print(f"\n   Agreements processed: {len(results)}/{len(common_agreements)}")
    
    if results:
        print(f"\n   Output files created:")
        for agreement in results.keys():
            print(f"     - {agreement}_matched.xlsx")
    
    return results


def run_matching(rate_card_file_path=None):
    """
    Run the matching workflow to match shipments with rate card.
    LEGACY: For backward compatibility with single rate card processing.
    
    For multi-agreement processing, use run_matching_all_agreements() instead.
    
    Args:
        rate_card_file_path (str, optional): Path to rate card file relative to "input/" folder.
                                            If None, will try to find rate_card.xlsx or rate_card.xls in input folder.
    
    Returns:
        str: Path to the output file (Matched_Shipments_with.xlsx) if successful, None otherwise
    """
    import sys
    
    print("="*80)
    print("RATE LANE MATCHING - Find Lane # from Rate Card for Shipments")
    print("="*80)
    
    # Step 1: Get rate card from part4_rate_card_processing.py
    print("\n1. Getting Rate Card from part4_rate_card_processing.py...")
    
    # If rate_card_file_path not provided, try to find it
    if rate_card_file_path is None:
        input_folder = "input"
        possible_names = ["rate_coty.xlsx", "rate_card.xls", "rate.xlsx"]
        for name in possible_names:
            full_path = os.path.join(input_folder, name)
            if os.path.exists(full_path):
                rate_card_file_path = name
                print(f"   Auto-detected rate card file: {rate_card_file_path}")
                break
        
        if rate_card_file_path is None:
            print(f"   [ERROR] Rate card file not found. Tried: {possible_names}")
            return None
    
    try:
        from part4_rate_card_processing import process_rate_card
        
        df_rate_card, rate_card_columns, rate_card_conditions, _ = process_rate_card(rate_card_file_path)
        
        print(f"   Rate Card loaded: {df_rate_card.shape[0]} rows x {df_rate_card.shape[1]} columns")
        print(f"   Rate Card columns: {len(rate_card_columns)}")
        print(f"   Conditions loaded: {len(rate_card_conditions)} columns with conditions")
        if rate_card_conditions:
            print(f"   Columns with conditions: {list(rate_card_conditions.keys())}")
        
    except ImportError as e:
        print(f"   [ERROR] Could not import part4_rate_card_processing: {e}")
        print("   Please ensure part4_rate_card_processing.py is in the same directory.")
        return None
    except Exception as e:
        print(f"   [ERROR] Failed to process rate card: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    # Step 1b: Load Business Rules for matching validation
    print("\n1b. Loading Business Rules for exact lane matching...")
    business_rules_lookup = None
    if BUSINESS_RULES_AVAILABLE:
        try:
            business_rules_lookup = load_business_rules_for_matching(rate_card_file_path)
            if business_rules_lookup:
                print(f"   Business Rules loaded successfully:")
                print(f"      - Rules with country: {len(business_rules_lookup.get('rule_to_country', {}))}")
                print(f"      - Rules with postal codes: {len(business_rules_lookup.get('rule_to_postal_codes', {}))}")
            else:
                print(f"   No business rules found (this is OK if the rate card doesn't have them)")
        except Exception as e:
            print(f"   [WARNING] Could not load business rules: {e}")
            print(f"   Matching will continue without business rule validation")
    
    # Step 2: Get dataframes from vocabular.py output (partly_df/vocabulary_mapping.xlsx)
    print("\n2. Loading Shipment dataframes from vocabular.py output...")
    
    # Vocabular output is stored in partly_df folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    vocabular_output_path = os.path.join(script_dir, "partly_df", "vocabulary_mapping.xlsx")
    
    if not os.path.exists(vocabular_output_path):
        print(f"   [ERROR] vocabulary_mapping.xlsx not found at: {vocabular_output_path}")
        print(f"   Please ensure vocabular.py has been run and the file exists in the partly_df folder.")
        return None
    
    print(f"   Found vocabulary_mapping.xlsx at: {vocabular_output_path}")

    
    try:
        # Read Excel file with all sheets
        excel_file = pd.ExcelFile(vocabular_output_path)
        sheet_names = excel_file.sheet_names
        print(f"   Found sheets in Excel file: {sheet_names}")
        
        etof_renamed = None
        lc_renamed = None
        origin_renamed = None
        
        # Read ETOF sheet if it exists
        if 'ETOF' in sheet_names:
            etof_renamed = pd.read_excel(vocabular_output_path, sheet_name='ETOF')
            print(f"   Loaded ETOF DataFrame: {etof_renamed.shape[0]} rows x {etof_renamed.shape[1]} columns")
        else:
            print(f"   [WARNING] ETOF sheet not found in vocabular_output.xlsx")
        
        # Read LC sheet if it exists
        if 'LC' in sheet_names:
            lc_renamed = pd.read_excel(vocabular_output_path, sheet_name='LC')
            print(f"   Loaded LC DataFrame: {lc_renamed.shape[0]} rows x {lc_renamed.shape[1]} columns")
        else:
            print(f"   [WARNING] LC sheet not found in vocabular_output.xlsx")
        
        # Read Origin sheet if it exists
        if 'Origin' in sheet_names:
            origin_renamed = pd.read_excel(vocabular_output_path, sheet_name='Origin')
            print(f"   Loaded Origin DataFrame: {origin_renamed.shape[0]} rows x {origin_renamed.shape[1]} columns")
        else:
            print(f"   [INFO] Origin sheet not found in vocabular_output.xlsx (optional)")
        
        if etof_renamed is None and lc_renamed is None:
            print(f"   [ERROR] No ETOF or LC dataframes found in vocabulary_mapping.xlsx")
            print(f"   Please ensure vocabular.py has been run and generated the Excel file with ETOF or LC sheets.")
            return None
    
    except FileNotFoundError:
        print(f"   [ERROR] File not found: {vocabular_output_path}")
        print(f"   Please run vocabular.py first to generate partly_df/vocabulary_mapping.xlsx")
        return None
    except Exception as e:
        print(f"   [ERROR] Failed to read vocabular_output.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    # Step 3: Select dataframe to process (LC if present, otherwise ETOF)
    print("\n3. Selecting shipment dataframe:")
    
    df_to_process = None
    shipment_type = None
    
    # Priority: LC first, then ETOF
    if lc_renamed is not None and not lc_renamed.empty:
        df_to_process = lc_renamed
        shipment_type = "LC"
        print(f"   Using LC DataFrame: {df_to_process.shape[0]} rows x {df_to_process.shape[1]} columns")
    elif etof_renamed is not None and not etof_renamed.empty:
        df_to_process = etof_renamed
        shipment_type = "ETOF"
        print(f"   LC not available, using ETOF DataFrame: {df_to_process.shape[0]} rows x {df_to_process.shape[1]} columns")
    else:
        print("\n   [ERROR] No LC or ETOF dataframes available to process!")
        print("   Please ensure vocabular.py has been run and generated LC or ETOF sheets.")
        sys.exit(1)
    
    # Step 4: Filter to only rows with values in ETOF # column
    print(f"\n4. Filtering rows with values in ETOF # column...")
    
    # Find ETOF # column (handle variations)
    etof_col = None
    etof_col_variations = ['ETOF #', 'ETOF#', 'etof #', 'etof#', 'ETOF', 'etof']
    
    for col in df_to_process.columns:
        col_normalized = str(col).strip()
        for variation in etof_col_variations:
            if col_normalized.lower() == variation.lower() or col_normalized.lower().replace(' ', '') == variation.lower().replace(' ', ''):
                etof_col = col
                break
        if etof_col:
            break
    
    if etof_col:
        print(f"   Found ETOF column: '{etof_col}'")
        initial_row_count = len(df_to_process)
        
        # Filter to keep only rows where ETOF # has a value (not null, not empty, not NaN)
        df_to_process = df_to_process[df_to_process[etof_col].notna()]
        df_to_process = df_to_process[df_to_process[etof_col].astype(str).str.strip() != '']
        df_to_process = df_to_process[df_to_process[etof_col].astype(str).str.lower() != 'nan']
        
        filtered_row_count = len(df_to_process)
        removed_rows = initial_row_count - filtered_row_count
        
        print(f"   Initial rows: {initial_row_count}")
        print(f"   Rows with ETOF # values: {filtered_row_count}")
        print(f"   Rows removed (no ETOF # value): {removed_rows}")
        
        if filtered_row_count == 0:
            print(f"\n   [ERROR] No rows remaining after filtering for ETOF # values!")
            print(f"   Please ensure the dataframe has rows with values in the ETOF # column.")
            return None
    else:
        print(f"   [WARNING] ETOF # column not found. Processing all rows.")
        print(f"   Searched for columns: {etof_col_variations}")
        print(f"   Available columns: {list(df_to_process.columns)}")
    
    # Step 5: Print input dataframe before matching
    print(f"\n5. Input {shipment_type} DataFrame before matching:")
    print(f"   Shape: {df_to_process.shape[0]} rows x {df_to_process.shape[1]} columns")
    print(f"   Columns: {list(df_to_process.columns)}")
    print(f"\n   First few rows:")
    print(df_to_process.head())
    
    # Step 6: Find common columns and match
    print(f"\n6. Finding common columns...")
    common_columns = find_common_columns(df_to_process, df_rate_card)
    
    if not common_columns:
        print("\nError: No common columns found between shipment and rate card dataframes.")
        return None
    
    # Step 7: Match shipments with rate card to find Lane #
    print("\n" + "="*80)
    print("FINDING RATE LANES FOR SHIPMENTS")
    print("="*80)
    print("Note: Business rules (postal code zones, country regions) will be validated for exact lane matching.")
    
    df_result = match_shipments_with_rate_card(df_to_process, df_rate_card, common_columns, business_rules_lookup=business_rules_lookup)
    
    # Step 8: Reorder columns and save results
    print("\n8. Reordering columns and saving results...")
    # Handle Colab environment where __file__ is not defined
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        # In Colab or interactive environments, use current working directory
        script_dir = os.getcwd()
    
    # Use absolute path to ensure it works even after directory changes
    output_file = os.path.abspath(os.path.join(script_dir, "Matched_Shipments_with.xlsx"))
    print(f"   Output file will be saved to: {output_file}")
    
    # Reorder columns: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date, then others
    def reorder_columns(df):
        """Reorder columns with priority: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date, then others."""
        if df is None or df.empty:
            return df
        
        # Find priority columns (handle variations)
        priority_columns = []
        other_columns = []
        
        # Define priority column patterns
        lc_patterns = ['LC #', 'LC#', 'lc #', 'lc#']
        etof_patterns = ['ETOF #', 'ETOF#', 'etof #', 'etof#']
        shipment_id_patterns = ['Shipment ID', 'ShipmentID', 'shipment id', 'shipmentid', 
                               'SHIPMENT_ID', 'SHIPMENT ID', 'Shipment', 'shipment', 'SHIPMENT']
        delivery_patterns = ['Delivery Number', 'DeliveryNumber', 'delivery number', 'deliverynumber',
                           'DELIVERY_NUMBER', 'Delivery', 'delivery', 'DELIVERY']
        carrier_patterns = ['Carrier', 'carrier', 'CARRIER']
        ship_date_patterns = ['SHIP_DATE', 'ship_date', 'Ship Date', 'ship date', 'SHIP DATE',
                             'Loading date', 'Loading Date', 'loading date', 'LOADING DATE']
        
        # Find and collect priority columns
        lc_col = None
        etof_col = None
        shipment_id_col = None
        delivery_col = None
        carrier_col = None
        ship_date_col = None
        
        for col in df.columns:
            col_str = str(col).strip()
            col_lower = col_str.lower().replace(' ', '').replace('_', '')
            
            # Check for LC #
            if not lc_col:
                for pattern in lc_patterns:
                    if col_lower == pattern.lower().replace(' ', '') or col_str == pattern:
                        lc_col = col
                        break
            
            # Check for ETOF #
            if not etof_col:
                for pattern in etof_patterns:
                    if col_lower == pattern.lower().replace(' ', '') or col_str == pattern:
                        etof_col = col
                        break
            
            # Check for Shipment ID
            if not shipment_id_col:
                for pattern in shipment_id_patterns:
                    pattern_lower = pattern.lower().replace(' ', '').replace('_', '')
                    if col_lower == pattern_lower or col_str == pattern:
                        shipment_id_col = col
                        break
                # Also check if column contains "shipment" and "id"
                if not shipment_id_col and 'shipment' in col_lower and 'id' in col_lower:
                    shipment_id_col = col
            
            # Check for Delivery Number
            if not delivery_col:
                for pattern in delivery_patterns:
                    pattern_lower = pattern.lower().replace(' ', '').replace('_', '')
                    if col_lower == pattern_lower or col_str == pattern:
                        delivery_col = col
                        break
                # Also check if column contains "delivery" and "number"
                if not delivery_col and 'delivery' in col_lower and 'number' in col_lower:
                    delivery_col = col
            
            # Check for Carrier
            if not carrier_col:
                for pattern in carrier_patterns:
                    if col_lower == pattern.lower() or col_str == pattern:
                        carrier_col = col
                        break
            
            # Check for Ship date
            if not ship_date_col:
                for pattern in ship_date_patterns:
                    if col_lower == pattern.lower().replace(' ', '_') or col_str == pattern:
                        ship_date_col = col
                        break
        
        # Build ordered column list
        ordered_columns = []
        
        # Add priority columns in order: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date
        if lc_col:
            ordered_columns.append(lc_col)
        if etof_col:
            ordered_columns.append(etof_col)
        if shipment_id_col:
            ordered_columns.append(shipment_id_col)
        if delivery_col:
            ordered_columns.append(delivery_col)
        if carrier_col:
            ordered_columns.append(carrier_col)
        if ship_date_col:
            ordered_columns.append(ship_date_col)
        
        # Add all other columns (excluding priority columns and comment)
        priority_set = {lc_col, etof_col, shipment_id_col, delivery_col, carrier_col, ship_date_col}
        for col in df.columns:
            if col not in priority_set and col != 'comment':
                ordered_columns.append(col)
        
        # Add comment column last
        if 'comment' in df.columns:
            ordered_columns.append('comment')
        
        # Reorder dataframe
        df_reordered = df[ordered_columns]
        
        print(f"   Column order: {ordered_columns[:7]}... (and {len(ordered_columns) - 7} more)")
        
        return df_reordered
    
    # Reorder result dataframe
    df_result_reordered = reorder_columns(df_result)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        def format_sheet(ws, df_for_sheet):
            """Apply formatting to a worksheet."""
            # Style header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width with some padding, but cap at 50
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Style comment column if it exists
            if 'comment' in df_for_sheet.columns:
                comment_col_idx = list(df_for_sheet.columns).index('comment') + 1
                comment_col_letter = get_column_letter(comment_col_idx)
                
                # Make comment column wider
                ws.column_dimensions[comment_col_letter].width = 60
                
                # Wrap text in comment column
                for row in ws.iter_rows(min_row=2, min_col=comment_col_idx, max_col=comment_col_idx):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Find 'Carrier agreement #' column (case-insensitive)
        carrier_agreement_col = None
        for col in df_result_reordered.columns:
            if 'carrier agreement' in str(col).lower():
                carrier_agreement_col = col
                break
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if carrier_agreement_col and df_result_reordered[carrier_agreement_col].notna().any():
                # Group by Carrier agreement # and create separate sheets
                unique_agreements = df_result_reordered[carrier_agreement_col].dropna().unique()
                print(f"\n   Found {len(unique_agreements)} unique Carrier agreement # values")
                
                for agreement in unique_agreements:
                    # Filter data for this agreement
                    df_agreement = df_result_reordered[df_result_reordered[carrier_agreement_col] == agreement]
                    
                    # Create sheet name (Excel limits to 31 chars, remove invalid chars)
                    sheet_name = str(agreement)[:31]
                    # Remove invalid Excel sheet name characters
                    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
                    for char in invalid_chars:
                        sheet_name = sheet_name.replace(char, '_')
                    
                    df_agreement.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"   Created sheet '{sheet_name}' with {len(df_agreement)} rows")
                
                # Add rows without Carrier agreement # to a separate sheet
                df_no_agreement = df_result_reordered[df_result_reordered[carrier_agreement_col].isna()]
                if not df_no_agreement.empty:
                    df_no_agreement.to_excel(writer, sheet_name='No Agreement', index=False)
                    print(f"   Created sheet 'No Agreement' with {len(df_no_agreement)} rows")
                
                # Apply formatting to all sheets
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    format_sheet(ws, df_result_reordered)
            else:
                # No Carrier agreement # column found, save as single sheet
                print(f"\n   [INFO] 'Carrier agreement #' column not found, saving as single sheet")
                df_result_reordered.to_excel(writer, sheet_name='Matched Shipments', index=False)
                
                # Apply formatting
                workbook = writer.book
                if 'Matched Shipments' in workbook.sheetnames:
                    ws = workbook['Matched Shipments']
                    format_sheet(ws, df_result_reordered)
        
        print(f"\n[SUCCESS] Results saved to: {output_file}")
        if carrier_agreement_col:
            print(f"  - Sheets created per Carrier agreement # (formatted)")
        else:
            print(f"  - Sheet: Matched Shipments with Rate Lanes (formatted)")
        print(f"\nTotal rows processed: {len(df_result)}")
        print(f"Total columns: {len(df_result.columns)} (reordered: LC #, ETOF #, Carrier, Ship date, then others)")
        
        # Show summary
        rows_with_lane = df_result[df_result['comment'].str.startswith('Rate lane', na=False)]
        rows_no_match = df_result[df_result['comment'] == 'No matching rate card entries found']
        print(f"  - Rows with Rate Lane assigned: {len(rows_with_lane)}")
        print(f"  - Rows with no matching rate card: {len(rows_no_match)}")
        
    except ImportError:
        # Fallback if openpyxl formatting is not available
        print("   [WARNING] openpyxl formatting not available, saving without formatting...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_result_reordered.to_excel(writer, sheet_name='Matched Shipments', index=False)
        print(f"\n[SUCCESS] Results saved to: {output_file} (without formatting)")
    except PermissionError:
        print(f"\n[ERROR] Permission denied: Cannot write to {output_file}")
        print("   The file is likely open in Excel. Please close it and run again.")
        return None
    except Exception as e:
        print(f"\n[ERROR] Failed to save results: {e}")
        import traceback
        traceback.print_exc()
        
        return None
    
    print(f"\n✅ Rate lane matching complete! Results saved to: {output_file}")
    print("="*80)
    
    return output_file


def create_lc_etof_with_comments():
    """
    Create a copy of lc_etof_mapping.xlsx with comments added from matched files.
    
    This function:
    1. Reads lc_etof_mapping.xlsx
    2. For each agreement tab, finds the corresponding <agreement>_matched.xlsx file
    3. Adds/updates the 'comment' column by matching on ETOF #
    4. Updates the 'All Data' tab with all comments
    5. Saves as lc_etof_with_comments.xlsx
    
    Returns:
        str: Path to output file if successful, None otherwise
    """
    print("\n" + "="*80)
    print("CREATING LC_ETOF WITH COMMENTS")
    print("="*80)
    
    partly_df = get_partly_df_folder()
    
    # Step 1: Find lc_etof_mapping.xlsx
    print("\n1. Finding lc_etof_mapping.xlsx...")
    mapping_file = get_lc_etof_mapping_file()
    
    if mapping_file is None:
        print("   [ERROR] lc_etof_mapping.xlsx not found in partly_df/")
        return None
    
    print(f"   Found: {mapping_file}")
    
    # Step 2: Read all sheets from the mapping file
    print("\n2. Reading all sheets...")
    try:
        xl = pd.ExcelFile(mapping_file)
        all_sheets = {}
        
        for sheet_name in xl.sheet_names:
            all_sheets[sheet_name] = pd.read_excel(mapping_file, sheet_name=sheet_name)
            print(f"   - {sheet_name}: {len(all_sheets[sheet_name])} rows")
        
    except Exception as e:
        print(f"   [ERROR] Could not read lc_etof_mapping.xlsx: {e}")
        return None
    
    # Step 3: Find all matched files
    print("\n3. Finding matched files...")
    matched_files = {}
    for file in partly_df.glob("*_matched.xlsx"):
        # Extract agreement number (e.g., "RA20220420022_matched.xlsx" -> "RA20220420022")
        agreement = file.stem.replace("_matched", "")
        matched_files[agreement] = file
        print(f"   - {agreement}: {file.name}")
    
    # Also check for Matched_Shipments_with.xlsx (created by run_matching)
    matched_shipments_file = partly_df.parent / "Matched_Shipments_with.xlsx"
    if not matched_shipments_file.exists():
        matched_shipments_file = partly_df / "Matched_Shipments_with.xlsx"
    
    if not matched_files and not matched_shipments_file.exists():
        print("   [WARNING] No matched files found. Run matching first.")
        return None
    
    # If no individual matched files but Matched_Shipments_with.xlsx exists, use it
    if not matched_files and matched_shipments_file.exists():
        print(f"   Using combined file: {matched_shipments_file.name}")
        matched_files['_combined'] = matched_shipments_file
    
    # Step 4: Build ETOF # to comment mapping from all matched files
    print("\n4. Building ETOF # to comment mapping...")
    etof_to_comment = {}  # {etof_number: comment}
    agreement_etof_comments = {}  # {agreement: {etof_number: comment}}
    
    etof_variations = ['ETOF #', 'ETOF#', 'etof #', 'etof#', 'ETOF', 'etof']
    
    for agreement, matched_file in matched_files.items():
        try:
            # Handle combined file (Matched_Shipments_with.xlsx) vs individual files
            if agreement == '_combined':
                # Read all sheets from the combined file
                xl = pd.ExcelFile(matched_file)
                for sheet_name in xl.sheet_names:
                    if sheet_name.lower() in ['no agreement', 'sheet1']:
                        continue  # Skip non-agreement sheets
                    
                    df_matched = pd.read_excel(matched_file, sheet_name=sheet_name)
                    
                    # Find ETOF # column
                    etof_col = None
                    for col in df_matched.columns:
                        for var in etof_variations:
                            if col.lower().replace(' ', '') == var.lower().replace(' ', ''):
                                etof_col = col
                                break
                        if etof_col:
                            break
                    
                    if not etof_col:
                        continue
                    
                    # Check if 'comment' column exists
                    if 'comment' not in df_matched.columns:
                        continue
                    
                    if sheet_name not in agreement_etof_comments:
                        agreement_etof_comments[sheet_name] = {}
                    
                    for _, row in df_matched.iterrows():
                        etof_val = row.get(etof_col)
                        comment_val = row.get('comment')
                        
                        if pd.notna(etof_val) and str(etof_val).strip() and str(etof_val).lower() != 'nan':
                            etof_key = str(etof_val).strip()
                            if pd.notna(comment_val):
                                comment_str = str(comment_val)
                                etof_to_comment[etof_key] = comment_str
                                agreement_etof_comments[sheet_name][etof_key] = comment_str
                    
                    print(f"   - {sheet_name}: {len(agreement_etof_comments[sheet_name])} comments loaded")
            else:
                # Standard individual matched file
                df_matched = pd.read_excel(matched_file, sheet_name='Matched Shipments')
                
                # Find ETOF # column
                etof_col = None
                for col in df_matched.columns:
                    for var in etof_variations:
                        if col.lower().replace(' ', '') == var.lower().replace(' ', ''):
                            etof_col = col
                            break
                    if etof_col:
                        break
                
                if not etof_col:
                    print(f"   [WARNING] ETOF # column not found in {matched_file.name}")
                    continue
                
                # Check if 'comment' column exists
                if 'comment' not in df_matched.columns:
                    print(f"   [WARNING] 'comment' column not found in {matched_file.name}")
                    continue
                
                agreement_etof_comments[agreement] = {}
                
                for _, row in df_matched.iterrows():
                    etof_val = row.get(etof_col)
                    comment_val = row.get('comment')
                    
                    if pd.notna(etof_val) and str(etof_val).strip() and str(etof_val).lower() != 'nan':
                        etof_key = str(etof_val).strip()
                        if pd.notna(comment_val):
                            comment_str = str(comment_val)
                            etof_to_comment[etof_key] = comment_str
                            agreement_etof_comments[agreement][etof_key] = comment_str
                
                print(f"   - {agreement}: {len(agreement_etof_comments[agreement])} comments loaded")
            
        except Exception as e:
            print(f"   [ERROR] Could not read {matched_file.name}: {e}")
    
    print(f"\n   Total unique ETOF # with comments: {len(etof_to_comment)}")
    
    # Step 5: Update each sheet with comments
    print("\n5. Updating sheets with comments...")
    
    for sheet_name, df_sheet in all_sheets.items():
        # Find ETOF # column in this sheet
        etof_col_sheet = None
        for col in df_sheet.columns:
            for var in etof_variations:
                if col.lower().replace(' ', '') == var.lower().replace(' ', ''):
                    etof_col_sheet = col
                    break
            if etof_col_sheet:
                break
        
        if not etof_col_sheet:
            print(f"   - {sheet_name}: ETOF # column not found, skipping")
            continue
        
        # Add/update comment column
        if 'comment' not in df_sheet.columns:
            df_sheet['comment'] = None
        
        # Determine which comments to use
        if sheet_name in agreement_etof_comments:
            # Use agreement-specific comments
            comments_to_use = agreement_etof_comments[sheet_name]
        else:
            # Use all comments (for "All Data" and other tabs)
            comments_to_use = etof_to_comment
        
        # Update comments based on ETOF #
        updated_count = 0
        for idx, row in df_sheet.iterrows():
            etof_val = row.get(etof_col_sheet)
            if pd.notna(etof_val) and str(etof_val).strip() and str(etof_val).lower() != 'nan':
                etof_key = str(etof_val).strip()
                if etof_key in comments_to_use:
                    df_sheet.at[idx, 'comment'] = comments_to_use[etof_key]
                    updated_count += 1
        
        all_sheets[sheet_name] = df_sheet
        print(f"   - {sheet_name}: {updated_count} rows updated with comments")
    
    # Step 6: Save to new file
    print("\n6. Saving to lc_etof_with_comments.xlsx...")
    output_file = partly_df / "lc_etof_with_comments.xlsx"
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df_sheet in all_sheets.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\n   ✓ Saved to: {output_file}")
        
        # Show summary
        print("\n" + "="*80)
        print("SUMMARY")
        print("="*80)
        for sheet_name, df_sheet in all_sheets.items():
            if 'comment' in df_sheet.columns:
                has_comment = df_sheet['comment'].notna().sum()
                print(f"   {sheet_name}: {has_comment}/{len(df_sheet)} rows have comments")
        
        return str(output_file)
        
    except PermissionError:
        print(f"   [ERROR] Permission denied. Close the file if it's open in Excel.")
        return None
    except Exception as e:
        print(f"   [ERROR] Failed to save: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    # Initialize logger to capture output to text file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_path = f"partly_df/matching_output_{timestamp}.txt"
    
    # Ensure partly_df directory exists
    os.makedirs("partly_df", exist_ok=True)
    
    with init_logger(log_file_path) as logger:
        # Redirect stdout to our logger
        old_stdout = sys.stdout
        sys.stdout = logger
        
        try:
            print("\n" + "="*80)
            print("RATE LANE MATCHING")
            print("="*80)
            print(f"\nLog file: {log_file_path}")
            print("\nThis script matches shipments with rate cards to find Lane #.")
            print("\nExpected input files in 'partly_df/' folder:")
            print("  1. <agreement>_vocabulary_mapping.xlsx - created by vocabular.py")
            print("     (e.g., 'RA20220420022_vocabulary_mapping.xlsx')")
            print("  2. <agreement>.xlsx - rate cards created by part4_rate_card_processing.py")
            print("     (e.g., 'RA20220420022.xlsx')")
            print("  3. lc_etof_mapping.xlsx - created by part7_optional_order_lc_etof_mapping.py")
            print("\nOutput files:")
            print("  - <agreement>_matched.xlsx (per agreement)")
            print("  - lc_etof_with_comments.xlsx (combined with comments)")
            print(f"  - {log_file_path} (this log file)")
            
            # Run matching for all agreements
            results = run_matching_all_agreements()
            
            if results:
                print("\n" + "="*80)
                print("MATCHING SUCCESS!")
                print("="*80)
                for agreement, file_path in results.items():
                    print(f"\n  {agreement}:")
                    print(f"    - Matched: {file_path}")
                
                # Create lc_etof_with_comments.xlsx
                comments_file = create_lc_etof_with_comments()
                if comments_file:
                    print(f"\n  Combined file with comments:")
                    print(f"    - {comments_file}")
            else:
                print("\n[WARNING] No output files created. Please check error messages above.")
            
            print(f"\n\nFull output saved to: {log_file_path}")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    print(f"\n[INFO] Full matching output saved to: {log_file_path}")
