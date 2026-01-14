import pandas as pd
import openpyxl
import os


def process_rate_card(file_path):
    """
    Process a Rate Card Excel file from the input folder.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder (e.g., "rate_card.xlsx")
    
    Returns:
        tuple: (dataframe, list of column names, conditions dictionary, agreement number)
            - dataframe: Processed pandas DataFrame (filtered to black font columns)
            - list: List of column names in the processed dataframe
            - dict: Dictionary of conditions where keys are column names and values are condition text
            - str: Agreement number from "General Info" tab
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Extract Agreement number from "General Info" tab
    agreement_number = None
    try:
        workbook_info = openpyxl.load_workbook(full_path, data_only=True)
        if "General info" in workbook_info.sheetnames:
            general_info_sheet = workbook_info["General info"]
            # Find row with "Agreement number" in column A

            for row in general_info_sheet.iter_rows(min_col=1, max_col=2):
                print(row)
                cell_a = row[0]
                print(cell_a.value)
                cell_b = row[1] if len(row) > 1 else None
                print(cell_b.value)
                if cell_a.value and "Agreement number" in str(cell_a.value):
                    if cell_b and cell_b.value:
                        agreement_number = str(cell_b.value).strip()
                    break
        workbook_info.close()
    except Exception as e:
        print(f"Warning: Could not extract Agreement number from General Info tab: {e}")
    
    if agreement_number:
        print(f"   Agreement number: {agreement_number}")
    
    # Read the Excel file
    df_rate_card = pd.read_excel(full_path, sheet_name="Rate card", skiprows=2)
    
    # Find first column index (where data actually starts)
    first_column_index = None
    if df_rate_card is not None:
        for i, col in enumerate(df_rate_card.columns):
            if "nan" not in str(df_rate_card.iloc[0, i]).lower():
                first_column_index = i
                break
    
    if first_column_index is not None:
        df_rate_card = df_rate_card.iloc[:, :first_column_index]
    
    # Drop rows where the first column is NaN
    if df_rate_card is not None:
        df_rate_card.dropna(subset=[df_rate_card.columns[0]], inplace=True)
    
    # Set column names from first row
    new_columns = df_rate_card.iloc[0].tolist()
    df_rate_card.columns = new_columns
    df_rate_card = df_rate_card.iloc[1:]
    
    # Load the workbook to extract conditions and check font colors
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    sheet = workbook["Rate card"]
    
    # Find the header row that contains "Currency"
    first_data_row_index = None
    currency_index = None
    
    for row_index in range(1, min(10, sheet.max_row + 1)):
        row = sheet[row_index]
        row_values = [cell.value for cell in row]
        if "Currency" in row_values:
            currency_index = row_values.index("Currency")
            first_data_row_index = row_index
            break
    
    black_font_values = []
    column_notes = {}  # Will store conditions/notes for each column
    
    if first_data_row_index is not None and currency_index is not None:
        # Access the data in this row
        first_data_row = sheet[first_data_row_index]
        first_data_values = [cell.value for cell in first_data_row]
        truncated_data_values = first_data_values[:currency_index]
        
        # Extract conditional rules/notes from multiple sources:
        # 1. Comments (notes) in the header row cells
        # 2. Cell values in the row ABOVE the header (row above column name)
        # 3. Cell values in row 2 (legacy fallback)
        header_row_index = first_data_row_index
        if header_row_index and header_row_index <= sheet.max_row:
            for i, col_name in enumerate(truncated_data_values, 1):
                if col_name:  # Only process non-empty column names
                    header_cell = sheet.cell(row=header_row_index, column=i)
                    
                    # Source 1: Check for comments (where conditional rules are stored)
                    if header_cell.comment:
                        comment_text = header_cell.comment.text
                        if comment_text and comment_text.strip():
                            column_notes[col_name] = comment_text.strip()
                    
                    # Source 2: Check the cell ABOVE the column name header
                    if col_name not in column_notes:
                        above_row_index = header_row_index - 1
                        if above_row_index >= 1:
                            above_cell = sheet.cell(row=above_row_index, column=i)
                            if above_cell.value and str(above_cell.value).strip():
                                column_notes[col_name] = str(above_cell.value).strip()
                    
                    # Source 3: Also check for cell value notes in row 2 (legacy fallback)
                    if col_name not in column_notes:
                        notes_row_index = 2
                        if notes_row_index <= sheet.max_row and notes_row_index != header_row_index - 1:
                            note_cell = sheet.cell(row=notes_row_index, column=i)
                            if note_cell.value and str(note_cell.value).strip():
                                column_notes[col_name] = str(note_cell.value).strip()
        
        # Check font color to identify black font columns (required columns)
        for i, value in enumerate(truncated_data_values):
            if i < len(first_data_row):
                cell = first_data_row[i]
                font_color = "black"
                if cell.font and cell.font.color:
                    hex_color = cell.font.color.rgb
                    if hex_color is not None:
                        # Convert to string and handle different formats
                        hex_str = str(hex_color).upper()
                        # Remove 'FF' prefix if present (ARGB format)
                        if hex_str.startswith('FF') and len(hex_str) == 8:
                            hex_str = hex_str[2:]
                        
                        # Check if it's black
                        if hex_str == '000000' or hex_str == '00000000':
                            font_color = "black"
                        else:
                            # Check if it's a shade of grey (R, G, and B are close)
                            try:
                                if len(hex_str) == 6:
                                    r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
                                    # Check if it's a shade of grey (R, G, and B are close)
                                    if abs(r - g) < 10 and abs(g - b) < 10 and r > 0:  # Grey (not black, not white)
                                        font_color = "grey"
                                    else:
                                        font_color = "other non-black"  # For colors that are not black or grey
                            except (ValueError, IndexError):
                                pass
                
                if font_color == "black":
                    black_font_values.append(value)
    
    # Filter the DataFrame to keep only the columns whose names are in black_font_values
    if df_rate_card is not None and black_font_values:
        # Only include columns that actually exist in the dataframe
        available_columns = [col for col in black_font_values if col in df_rate_card.columns]
        if available_columns:
            df_filtered_rate_card = df_rate_card[available_columns]
        else:
            df_filtered_rate_card = df_rate_card
    else:
        df_filtered_rate_card = df_rate_card
    
    # Get list of column names
    column_names = df_filtered_rate_card.columns.tolist()
    
    # Create conditions dictionary (only for columns that exist in the filtered dataframe)
    conditions = {}
    for col_name in column_names:
        if col_name in column_notes:
            conditions[col_name] = column_notes[col_name]
    
    return df_filtered_rate_card, column_names, conditions, agreement_number


def clean_condition_text(condition_text):
    """
    Clean up condition text for better readability.
    
    Transforms:
        "Conditional rules:
        1. 33321-6422: TOPOSTALCODE starts with 33321-6422,333216422"
    To:
        "1. 33321-6422: starts with 33321-6422,333216422"
    """
    import re
    
    if not condition_text:
        return condition_text
    
    # Remove "Conditional rules:" header (case insensitive)
    cleaned = re.sub(r'(?i)^conditional\s*rules\s*:\s*\n?', '', condition_text.strip())
    
    # Remove column name references like "TOPOSTALCODE ", "FROMPOSTALCODE ", etc.
    # Pattern: After the colon and value identifier, remove uppercase column names followed by space
    # Example: "33321-6422: TOPOSTALCODE starts with" -> "33321-6422: starts with"
    cleaned = re.sub(r':\s*[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r': \1', cleaned)
    
    # Also handle cases without numbered format
    cleaned = re.sub(r'^[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r'\1', cleaned, flags=re.MULTILINE)
    
    # Clean up extra whitespace and newlines
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    cleaned = '\n'.join(lines)
    
    return cleaned


# =============================================================================
# BUSINESS RULES PROCESSING
# =============================================================================

def process_business_rules(file_path):
    """
    Process the Business rules tab from a Rate Card Excel file.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder
    
    Returns:
        dict: Dictionary containing:
            - 'postal_code_zones': list of zone rules with name, country, postal_codes, exclude
            - 'country_regions': list of region rules with name, country, postal_codes, exclude
            - 'no_data_added': list of entries with no data
            - 'raw_rules': all parsed rules as a list of dicts
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Load the workbook
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    
    # Check if "Business rules" sheet exists
    if "Business rules" not in workbook.sheetnames:
        print(f"   [WARNING] 'Business rules' sheet not found in {file_path}")
        return {
            'postal_code_zones': [],
            'country_regions': [],
            'no_data_added': [],
            'raw_rules': []
        }
    
    sheet = workbook["Business rules"]
    
    # DEBUG: Print sheet info
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SHEET ANALYSIS")
    print(f"{'='*60}")
    print(f"   Sheet name: 'Business rules'")
    print(f"   Total rows in sheet: {sheet.max_row}")
    print(f"   Total columns: {sheet.max_column}")
    
    # STEP 1: Read all rows and filter out empty ones (skip first 2 rows)
    print(f"\n   [DEBUG] Step 1: Reading and filtering rows (skipping first 2 rows)...")
    
    all_rows = []  # Will store (original_row_idx, row_values) tuples
    for row_idx in range(3, sheet.max_row + 1):
        row = sheet[row_idx]
        row_values = [cell.value for cell in row]
        
        # Check if row is empty
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == '') for v in row_values)
        
        if not is_empty:
            all_rows.append((row_idx, row_values))
    
    print(f"   [DEBUG] Total non-empty rows found: {len(all_rows)} (out of {sheet.max_row - 2} after skipping first 2)")
    
    # DEBUG: Print first 10 non-empty rows to see structure
    print(f"\n   [DEBUG] First 10 non-empty rows content:")
    for i, (row_idx, row_values) in enumerate(all_rows[:10]):
        non_empty = [(col_i, v) for col_i, v in enumerate(row_values) if v is not None]
        print(f"      Row {row_idx}: {non_empty}")
    
    if len(all_rows) > 10:
        print(f"      ... and {len(all_rows) - 10} more rows")
    
    # Marker values to look for (case-insensitive)
    markers = ['postal code zones', 'country regions', 'no data added']
    
    # Result structure
    result = {
        'postal_code_zones': [],
        'country_regions': [],
        'no_data_added': [],
        'raw_rules': []
    }
    
    # Track sections and their header columns
    current_section = None
    header_columns = {}  # Maps column index to header name
    waiting_for_header = False  # Flag to indicate we found a marker and are waiting for header row
    
    print(f"\n   [DEBUG] Step 2: Searching for markers: {markers}")
    print(f"   [DEBUG] Structure: MARKER row -> HEADER row (below) -> DATA rows")
    
    # Process non-empty rows
    for i, (row_idx, row_values) in enumerate(all_rows):
        # Check if this row contains a marker (section header)
        row_text_lower = ' '.join(str(v).lower() for v in row_values if v is not None)
        
        found_marker = None
        for marker in markers:
            if marker in row_text_lower:
                found_marker = marker
                print(f"\n   [DEBUG] >>> MARKER FOUND: '{marker}' at row {row_idx}")
                break
        
        if found_marker:
            # This is a marker row - next non-empty row will be the header
            current_section = found_marker.replace(' ', '_')
            waiting_for_header = True
            header_columns = {}  # Reset header columns for new section
            print(f"   [DEBUG]     Section: '{current_section}'")
            print(f"   [DEBUG]     Waiting for header row...")
            continue
        
        # If we're waiting for header, this row should be the header
        if waiting_for_header:
            waiting_for_header = False
            header_columns = {}
            
            print(f"   [DEBUG]     Header row (row {row_idx}): {[v for v in row_values if v is not None]}")
            
            for col_idx, cell_value in enumerate(row_values):
                if cell_value:
                    header_name = str(cell_value).strip().lower()
                    # Normalize header names
                    # IMPORTANT: Check 'exclude' BEFORE 'postal'/'code' to handle "Excluded Postal Code" columns
                    if 'name' in header_name:
                        header_columns[col_idx] = 'name'
                    elif 'country' in header_name:
                        header_columns[col_idx] = 'country'
                    elif 'exclude' in header_name:
                        # Must check before 'postal'/'code' since column might be "Excluded Postal Codes"
                        header_columns[col_idx] = 'exclude'
                    elif 'postal' in header_name or 'code' in header_name:
                        header_columns[col_idx] = 'postal_code'
                    else:
                        header_columns[col_idx] = header_name
            
            print(f"   [DEBUG]     Mapped header columns: {header_columns}")
            # Check if exclude column was found
            has_exclude = 'exclude' in header_columns.values()
            print(f"   [DEBUG]     Has EXCLUDE column: {has_exclude}")
            continue
        
        # If we're in a section and have header columns, parse the data row
        if current_section and header_columns:
            rule_data = {
                'section': current_section,
                'name': None,
                'country': None,
                'postal_code': None,
                'exclude': None
            }
            
            # Extract values based on header columns
            for col_idx, header_name in header_columns.items():
                if col_idx < len(row_values):
                    value = row_values[col_idx]
                    if value is not None:
                        rule_data[header_name] = str(value).strip() if value else None
            
            # Only add if we have at least a name or postal code
            if rule_data['name'] or rule_data['postal_code'] or rule_data['country']:
                # Debug: show rules with exclude values or DK rules
                if rule_data.get('exclude') or (rule_data.get('name') and 'DK' in str(rule_data.get('name', ''))):
                    print(f"   [DEBUG RAW] Rule: {rule_data.get('name')}, Postal: {str(rule_data.get('postal_code', ''))[:50]}, Exclude: {rule_data.get('exclude')}")
                result['raw_rules'].append(rule_data)
                
                # Add to appropriate section list
                if current_section == 'postal_code_zones':
                    result['postal_code_zones'].append(rule_data)
                elif current_section == 'country_regions':
                    result['country_regions'].append(rule_data)
                elif current_section == 'no_data_added':
                    result['no_data_added'].append(rule_data)
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SUMMARY")
    print(f"{'='*60}")
    print(f"   - Postal Code Zones: {len(result['postal_code_zones'])} rules")
    print(f"   - Country Regions: {len(result['country_regions'])} rules")
    print(f"   - No Data Added: {len(result['no_data_added'])} entries")
    print(f"   - Total raw rules: {len(result['raw_rules'])}")
    
    if not result['raw_rules']:
        print(f"\n   [WARNING] No rules were found! Possible issues:")
        print(f"      1. Markers not found in expected format")
        print(f"      2. Headers not in row above markers")
        print(f"      3. Data structure different than expected")
    
    return result


def transform_business_rules_to_conditions(business_rules):
    """
    Transform parsed business rules into condition format.
    Consolidates continuation rows (rows without a name) into the previous rule.
    
    Args:
        business_rules (dict): Output from process_business_rules()
    
    Returns:
        dict: Dictionary mapping zone/region names to their conditions
              Format: {zone_name: {'country': 'XX', 'postal_codes': ['12', '34'], 'excluded_postal_codes': [...]}}
    """
    conditions = {}
    
    # First pass: consolidate continuation rows into their parent rules
    consolidated_rules = []
    current_rule = None
    
    for rule in business_rules.get('raw_rules', []):
        name = rule.get('name')
        postal_code_str = rule.get('postal_code', '')
        exclude_str = rule.get('exclude', '')
        country = rule.get('country')
        section = rule.get('section', '')
        
        if name:
            # This is a new rule - save the previous one if exists
            if current_rule:
                consolidated_rules.append(current_rule)
            
            current_rule = {
                'name': name,
                'section': section,
                'country': country,
                'postal_codes_list': [],
                'exclude_list': []
            }
            
            # Add postal codes from this row
            if postal_code_str:
                codes = [code.strip() for code in str(postal_code_str).split(',') if code.strip()]
                current_rule['postal_codes_list'].extend(codes)
            
            # Add excluded codes from this row
            if exclude_str:
                codes = [code.strip() for code in str(exclude_str).split(',') if code.strip()]
                current_rule['exclude_list'].extend(codes)
        
        elif current_rule:
            # This is a continuation row - add to current rule
            if postal_code_str:
                codes = [code.strip() for code in str(postal_code_str).split(',') if code.strip()]
                current_rule['postal_codes_list'].extend(codes)
            
            if exclude_str:
                codes = [code.strip() for code in str(exclude_str).split(',') if code.strip()]
                current_rule['exclude_list'].extend(codes)
            
            # Update country if provided in continuation row
            if country and not current_rule.get('country'):
                current_rule['country'] = country
    
    # Don't forget the last rule
    if current_rule:
        consolidated_rules.append(current_rule)
    
    print(f"\n[DEBUG] Consolidated {len(business_rules.get('raw_rules', []))} raw rows into {len(consolidated_rules)} rules")
    
    # Debug: Show rules with excluded codes
    rules_with_exclusions = [r for r in consolidated_rules if r.get('exclude_list')]
    if rules_with_exclusions:
        print(f"[DEBUG] Rules with EXCLUSIONS: {len(rules_with_exclusions)}")
        for rule in rules_with_exclusions[:10]:  # Show first 10
            print(f"   - '{rule['name']}': {len(rule['exclude_list'])} excluded codes: {rule['exclude_list'][:5]}{'...' if len(rule['exclude_list']) > 5 else ''}")
    else:
        print(f"[DEBUG] WARNING: No rules have exclusion codes!")
    
    # Debug: Show DK rules specifically
    dk_rules = [r for r in consolidated_rules if 'DK' in str(r.get('name', ''))]
    if dk_rules:
        print(f"\n[DEBUG] DK RULES ({len(dk_rules)} found):")
        for rule in dk_rules:
            print(f"   - '{rule['name']}': {len(rule.get('postal_codes_list', []))} postal, {len(rule.get('exclude_list', []))} excluded")
            if rule.get('exclude_list'):
                print(f"     Excluded: {rule['exclude_list'][:10]}...")
    
    # Second pass: transform consolidated rules into conditions format
    for rule in consolidated_rules:
        name = rule.get('name')
        if not name:
            continue
        
        section = rule.get('section', '')
        postal_codes = rule.get('postal_codes_list', [])
        excluded_postal_codes = rule.get('exclude_list', [])
        
        # Filter out non-numeric/invalid codes from excluded list (e.g., "yes", "true")
        excluded_postal_codes = [
            code for code in excluded_postal_codes 
            if code and any(c.isdigit() for c in code) and code.lower() not in ['yes', 'true', 'no', 'false']
        ]
        
        raw_postal_str = ', '.join(postal_codes) if postal_codes else ''
        raw_exclude_str = ', '.join(excluded_postal_codes) if excluded_postal_codes else ''
        
        condition = {
            'section': section,
            'country': rule.get('country'),
            'postal_codes': postal_codes,
            'excluded_postal_codes': excluded_postal_codes,
            'raw_postal_code': raw_postal_str,
            'raw_exclude': raw_exclude_str
        }
        
        conditions[name] = condition
        
        if excluded_postal_codes:
            print(f"   [DEBUG] Rule '{name}': {len(postal_codes)} postal codes, {len(excluded_postal_codes)} EXCLUDED codes")
    
    return conditions


def format_business_rule_condition(rule_name, condition):
    """
    Format a business rule condition into a readable string.
    
    Args:
        rule_name (str): Name of the rule/zone
        condition (dict): Condition dictionary from transform_business_rules_to_conditions
    
    Returns:
        str: Human-readable condition string
    """
    parts = []
    
    if condition.get('country'):
        parts.append(f"Country: {condition['country']}")
    
    if condition.get('postal_codes'):
        prefix_list = ', '.join(condition['postal_codes'][:5])
        if len(condition['postal_codes']) > 5:
            prefix_list += f", ... (+{len(condition['postal_codes']) - 5} more)"
        parts.append(f"Postal codes starting with: {prefix_list}")
    
    if condition.get('excluded_postal_codes'):
        exclude_list = ', '.join(condition['excluded_postal_codes'][:5])
        if len(condition['excluded_postal_codes']) > 5:
            exclude_list += f", ... (+{len(condition['excluded_postal_codes']) - 5} more)"
        parts.append(f"EXCLUDING: {exclude_list}")
    
    return ' | '.join(parts) if parts else 'No conditions'


def find_business_rule_columns(rate_card_df, business_rules_conditions):
    """
    Find which columns in the rate card contain business rule values.
    
    Args:
        rate_card_df (pd.DataFrame): The rate card dataframe
        business_rules_conditions (dict): Dictionary of business rule conditions with rule names as keys
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_columns': {rule_name: [list of columns where found]}
            - 'column_to_rules': {column_name: [list of rules found in it]}
            - 'unique_columns': set of unique column names that contain any business rule
    """
    rule_names = list(business_rules_conditions.keys())
    
    result = {
        'rule_to_columns': {},  # Which columns contain each rule
        'column_to_rules': {},  # Which rules are in each column
        'unique_columns': set()
    }
    
    if rate_card_df is None or rate_card_df.empty or not rule_names:
        return result
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] FINDING BUSINESS RULE COLUMNS IN RATE CARD")
    print(f"{'='*60}")
    print(f"   Searching for {len(rule_names)} rule names in {len(rate_card_df.columns)} columns...")
    
    # Columns to EXCLUDE from business rule detection (these contain codes, not business rule names)
    EXCLUDED_BUSINESS_RULE_COLUMNS = {
        'origin airport', 'destination airport', 'origin port', 'destination port',
        'pol', 'poe', 'port of loading', 'port of entry', 'airport', 'port',
        'origin airport code', 'destination airport code', 'airport code',
        'origin seaport', 'destination seaport', 'seaport',
        'ship_port', 'cust_port', 'origin_airport', 'destination_airport',
        'carrier', 'carrier name', 'carrier code', 'scac', 'scac code',
        'origin country', 'destination country', 'country', 'ship_country', 'cust_country'
    }
    
    # Create a set of rule names for faster lookup (case-insensitive)
    rule_names_lower = {str(name).lower(): name for name in rule_names}
    
    # For each column, check which rule names are present
    for col in rate_card_df.columns:
        # Skip excluded columns
        col_lower = str(col).lower().strip()
        if col_lower in EXCLUDED_BUSINESS_RULE_COLUMNS:
            print(f"   [SKIP] Column '{col}' excluded from business rule detection")
            continue
        try:
            # Get unique values in this column
            unique_values = rate_card_df[col].dropna().unique()
            
            # Check each unique value against rule names
            for val in unique_values:
                val_str = str(val).strip().lower()
                
                if val_str in rule_names_lower:
                    original_rule_name = rule_names_lower[val_str]
                    
                    # Track rule to columns mapping
                    if original_rule_name not in result['rule_to_columns']:
                        result['rule_to_columns'][original_rule_name] = []
                    if col not in result['rule_to_columns'][original_rule_name]:
                        result['rule_to_columns'][original_rule_name].append(col)
                    
                    # Track column to rules mapping
                    if col not in result['column_to_rules']:
                        result['column_to_rules'][col] = []
                    if original_rule_name not in result['column_to_rules'][col]:
                        result['column_to_rules'][col].append(original_rule_name)
                    
                    result['unique_columns'].add(col)
        except Exception as e:
            # Skip columns that can't be processed
            pass
    
    # Initialize empty lists for rules not found
    for rule_name in rule_names:
        if rule_name not in result['rule_to_columns']:
            result['rule_to_columns'][rule_name] = []
    
    # Print results
    print(f"\n   [RESULT] Unique columns containing business rules:")
    if result['unique_columns']:
        for col in sorted(result['unique_columns']):
            rules_in_col = result['column_to_rules'].get(col, [])
            print(f"      - '{col}': {len(rules_in_col)} rules found")
            # Show first few rules as examples
            if rules_in_col:
                examples = rules_in_col[:3]
                if len(rules_in_col) > 3:
                    print(f"         Examples: {examples} ... (+{len(rules_in_col) - 3} more)")
                else:
                    print(f"         Rules: {examples}")
    else:
        print(f"      No columns found containing business rule values")
    
    print(f"\n   [SUMMARY] {len(result['unique_columns'])} unique columns contain business rule values")
    
    return result


def get_business_rules_lookup(file_path):
    """
    Get a lookup dictionary from business rule names to their country and postal codes.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_country': {rule_name: country_code}
            - 'rule_to_postal_codes': {rule_name: [list of postal codes]}
            - 'rule_to_excluded_postal_codes': {rule_name: [list of excluded postal codes]}
            - 'business_rule_columns': set of column names containing business rules
            - 'all_rules': list of all rule data with name, country, postal_codes
    """
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Get rate card to find which columns contain business rules
    rate_card_df, rate_card_columns, _, _ = process_rate_card(file_path)
    business_rule_cols_info = find_business_rule_columns(rate_card_df, business_rules_conditions)
    
    result = {
        'rule_to_country': {},
        'rule_to_postal_codes': {},
        'rule_to_excluded_postal_codes': {},  # excluded postal codes
        'business_rule_columns': business_rule_cols_info.get('unique_columns', set()),
        'column_to_rules': business_rule_cols_info.get('column_to_rules', {}),
        'all_rules': []
    }
    
    for rule_name, condition in business_rules_conditions.items():
        country = condition.get('country')
        postal_codes = condition.get('postal_codes', [])
        excluded_postal_codes = condition.get('excluded_postal_codes', [])
        
        if country:
            result['rule_to_country'][rule_name] = country
        if postal_codes:
            result['rule_to_postal_codes'][rule_name] = postal_codes
        if excluded_postal_codes:
            result['rule_to_excluded_postal_codes'][rule_name] = excluded_postal_codes
        
        result['all_rules'].append({
            'name': rule_name,
            'country': country,
            'postal_codes': postal_codes,
            'excluded_postal_codes': excluded_postal_codes,
            'section': condition.get('section')
        })
    
    # Try to load excluded postal codes from the rate card Excel file if available
    # This is a fallback in case the input file parsing didn't capture exclusions
    try:
        # Construct path to the rate card Excel file in partly_df
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        rate_card_excel_path = os.path.join("partly_df", f"{base_name}.xlsx")
        
        print(f"\n[DEBUG] Looking for rate card Excel: {rate_card_excel_path} (from file_path: {file_path})")
        
        if os.path.exists(rate_card_excel_path):
            print(f"[DEBUG] Checking rate card Excel for excluded postal codes: {rate_card_excel_path}")
            xl = pd.ExcelFile(rate_card_excel_path)
            
            if 'Business Rules' in xl.sheet_names:
                df_br = pd.read_excel(xl, sheet_name='Business Rules')
                
                # Check if "Excluded Postal Codes" column exists
                if 'Excluded Postal Codes' in df_br.columns and 'Rule Name' in df_br.columns:
                    print(f"   [DEBUG] Found 'Excluded Postal Codes' column in Business Rules sheet")
                    
                    # Load excluded postal codes from Excel
                    excel_exclusions_loaded = 0
                    for _, row in df_br.iterrows():
                        rule_name = row.get('Rule Name')
                        excluded_str = row.get('Excluded Postal Codes', '')
                        
                        if rule_name and pd.notna(excluded_str) and str(excluded_str).strip():
                            # Parse the excluded postal codes (comma-separated)
                            excluded_codes = [code.strip() for code in str(excluded_str).split(',') if code.strip()]
                            
                            # Only add if we don't already have exclusions for this rule
                            if excluded_codes and rule_name not in result['rule_to_excluded_postal_codes']:
                                result['rule_to_excluded_postal_codes'][rule_name] = excluded_codes
                                excel_exclusions_loaded += 1
                            elif excluded_codes and rule_name in result['rule_to_excluded_postal_codes']:
                                # Merge with existing exclusions
                                existing = set(result['rule_to_excluded_postal_codes'][rule_name])
                                existing.update(excluded_codes)
                                result['rule_to_excluded_postal_codes'][rule_name] = list(existing)
                    
                    print(f"   [DEBUG] Loaded {excel_exclusions_loaded} rules with excluded postal codes from Excel")
                else:
                    print(f"   [DEBUG] 'Excluded Postal Codes' or 'Rule Name' column not found in Business Rules sheet")
                    print(f"   [DEBUG] Available columns: {list(df_br.columns)}")
            else:
                print(f"   [DEBUG] 'Business Rules' sheet not found in Excel file")
                print(f"   [DEBUG] Available sheets: {xl.sheet_names}")
        else:
            print(f"   [DEBUG] Rate card Excel file NOT FOUND: {rate_card_excel_path}")
    except Exception as e:
        print(f"   [WARNING] Could not load excluded postal codes from Excel: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"\n[DEBUG] Business Rules Lookup FINAL:")
    print(f"   - Rules with country: {len(result['rule_to_country'])}")
    print(f"   - Rules with postal codes: {len(result['rule_to_postal_codes'])}")
    print(f"   - Rules with excluded postal codes: {len(result['rule_to_excluded_postal_codes'])}")
    print(f"   - Columns containing rules: {sorted(result['business_rule_columns'])}")
    
    # Show some examples of excluded postal codes
    if result['rule_to_excluded_postal_codes']:
        print(f"\n[DEBUG] Sample excluded postal codes:")
        for rule_name, codes in list(result['rule_to_excluded_postal_codes'].items())[:5]:
            print(f"   - '{rule_name}': {len(codes)} codes - {codes[:5]}{'...' if len(codes) > 5 else ''}")
    
    return result


def get_required_geo_columns():
    """
    Get the list of required geographic columns that should be in the final output.
    These are derived from business rules and should be mapped from ETOF/LC files.
    
    Returns:
        list: List of required column names for origin/destination country and postal codes
    """
    return [
        'Origin Country',
        'Origin Postal Code', 
        'Destination Country',
        'Destination Postal Code'
    ]


def save_rate_card_output(file_path, output_path=None):
    """
    Process rate card and save output to Excel file with data and conditions.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
        output_path (str): Optional output path. If None, saves to "<agreement_number>.xlsx" in partly_df folder
    
    Returns:
        str: Path to the saved Excel file
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Process the rate card
    rate_card_dataframe, rate_card_column_names, rate_card_conditions, agreement_number = process_rate_card(file_path)
    
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Find which columns in rate card contain business rule values
    business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    # Set output path - save to partly_df folder (relative to script location)
    if output_path is None:
        # Get the directory where this script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Ensure partly_df folder exists in the script's directory
        partly_df_folder = os.path.join(script_dir, "partly_df")
        if not os.path.exists(partly_df_folder):
            os.makedirs(partly_df_folder)
        
        # Use agreement number as filename if available, otherwise use a default name
        if agreement_number:
            # Clean agreement number for use as filename (remove invalid characters)
            safe_agreement_number = "".join(c for c in agreement_number if c.isalnum() or c in ('-', '_', ' ')).strip()
            output_filename = f"{safe_agreement_number}.xlsx"
        else:
            # Fallback to original filename if no agreement number found
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_filename = f"{base_name}_processed.xlsx"
        
        output_path = os.path.join(partly_df_folder, output_filename)
    
    # Create conditions DataFrame with cleaned condition text
    conditions_data = []
    for col_name in rate_card_column_names:
        raw_condition = rate_card_conditions.get(col_name, "")
        cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
        conditions_data.append({
            'Column': col_name,
            'Has Condition': 'Yes' if col_name in rate_card_conditions else 'No',
            'Condition Rule': cleaned_condition
        })
    
    df_conditions = pd.DataFrame(conditions_data)
    
    # Save to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Rate Card Data
        rate_card_dataframe.to_excel(writer, sheet_name='Rate Card Data', index=False)
        
        # Sheet 2: Conditions
        df_conditions.to_excel(writer, sheet_name='Conditions', index=False)
        
        # Sheet 3: Business Rules
        business_rules_data = []
        for rule_name, condition in business_rules_conditions.items():
            # Get the columns where this rule is found
            rule_columns = business_rule_columns['rule_to_columns'].get(rule_name, [])
            columns_str = ', '.join(rule_columns) if rule_columns else '(not found in data)'
            
            # Get excluded postal codes
            excluded_postal_codes = condition.get('excluded_postal_codes', [])
            excluded_str = condition.get('raw_exclude', '')
            
            business_rules_data.append({
                'Rule Name': rule_name,
                'Section': condition.get('section', '').replace('_', ' ').title(),
                'Country': condition.get('country', ''),
                'Postal Codes': condition.get('raw_postal_code', ''),
                'Excluded Postal Codes': excluded_str,
                'Has Exclusions': 'Yes' if excluded_postal_codes else 'No',
                'Rate Card Columns': columns_str,
                'Formatted Condition': format_business_rule_condition(rule_name, condition)
            })
        
        df_business_rules = pd.DataFrame(business_rules_data)
        if not df_business_rules.empty:
            df_business_rules.to_excel(writer, sheet_name='Business Rules', index=False)
        
        # Sheet 4: Summary
        summary_data = {
            'Metric': [
                'Agreement Number',
                'Total Rows',
                'Total Columns',
                'Columns with Conditions',
                'Columns without Conditions',
                'Source File'
            ],
            'Value': [
                agreement_number if agreement_number else 'Not found',
                len(rate_card_dataframe),
                len(rate_card_column_names),
                len(rate_card_conditions),
                len(rate_card_column_names) - len(rate_card_conditions),
                file_path
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply formatting
        workbook = writer.book
        
        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        condition_yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        condition_no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format Rate Card Data sheet
        ws_data = workbook['Rate Card Data']
        for cell in ws_data[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_data.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        ws_data.freeze_panes = 'A2'
        
        # Format Conditions sheet
        ws_conditions = workbook['Conditions']
        for cell in ws_conditions[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color "Has Condition" column based on Yes/No
        for row in ws_conditions.iter_rows(min_row=2, max_row=ws_conditions.max_row):
            has_condition_cell = row[1]  # Column B (Has Condition)
            if has_condition_cell.value == 'Yes':
                has_condition_cell.fill = condition_yes_fill
            elif has_condition_cell.value == 'No':
                has_condition_cell.fill = condition_no_fill
            
            # Wrap text in Condition Rule column
            if len(row) > 2:
                row[2].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Set column widths for Conditions sheet
        ws_conditions.column_dimensions['A'].width = 30  # Column
        ws_conditions.column_dimensions['B'].width = 15  # Has Condition
        ws_conditions.column_dimensions['C'].width = 80  # Condition Rule
        
        ws_conditions.freeze_panes = 'A2'
        
        # Format Business Rules sheet (if it exists)
        if 'Business Rules' in workbook.sheetnames:
            ws_business_rules = workbook['Business Rules']
            for cell in ws_business_rules[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths for Business Rules sheet
            ws_business_rules.column_dimensions['A'].width = 25  # Rule Name
            ws_business_rules.column_dimensions['B'].width = 20  # Section
            ws_business_rules.column_dimensions['C'].width = 15  # Country
            ws_business_rules.column_dimensions['D'].width = 40  # Postal Codes
            ws_business_rules.column_dimensions['E'].width = 10  # Exclude
            ws_business_rules.column_dimensions['F'].width = 30  # Rate Card Columns
            ws_business_rules.column_dimensions['G'].width = 50  # Formatted Condition
            
            # Wrap text in Postal Codes and Formatted Condition columns
            for row in ws_business_rules.iter_rows(min_row=2, max_row=ws_business_rules.max_row):
                if len(row) > 3:
                    row[3].alignment = Alignment(wrap_text=True, vertical="top")  # Postal Codes
                if len(row) > 6:
                    row[6].alignment = Alignment(wrap_text=True, vertical="top")  # Formatted Condition
            
            ws_business_rules.freeze_panes = 'A2'
        
        # Format Summary sheet
        ws_summary = workbook['Summary']
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 60
    
    print(f"\n✅ Rate Card output saved to: {output_path}")
    print(f"   - Agreement Number: {agreement_number if agreement_number else 'Not found'}")
    print(f"   - Sheet 'Rate Card Data': {len(rate_card_dataframe)} rows x {len(rate_card_column_names)} columns")
    print(f"   - Sheet 'Conditions': {len(rate_card_conditions)} columns with conditions")
    print(f"   - Sheet 'Business Rules': {len(business_rules_conditions)} rules")
    print(f"   - Sheet 'Summary': Overview statistics")
    
    return output_path


def process_multiple_rate_cards(file_paths):
    """
    Process multiple rate card files and save each to a separate output file.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
                          (e.g., ["rate_card_1.xlsx", "rate_card_2.xlsx"])
    
    Returns:
        dict: Dictionary mapping agreement numbers to their output file paths
              {agreement_number: output_path, ...}
    """
    results = {}
    
    print(f"\n{'='*60}")
    print(f"Processing {len(file_paths)} rate card(s)...")
    print(f"{'='*60}")
    
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] Processing: {file_path}")
        print("-" * 40)
        
        try:
            output_path = save_rate_card_output(file_path)
            
            # Get agreement number for the results dictionary
            _, _, _, agreement_number = process_rate_card(file_path)
            
            if agreement_number:
                results[agreement_number] = output_path
            else:
                # Use filename as key if no agreement number
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                results[base_name] = output_path
                
        except Exception as e:
            print(f"❌ Error processing {file_path}: {e}")
            continue
    
    print(f"\n{'='*60}")
    print(f"Processing complete! {len(results)}/{len(file_paths)} files processed successfully.")
    print(f"{'='*60}")
    
    if results:
        print("\nOutput files created:")
        for agreement, path in results.items():
            print(f"  - {agreement}: {path}")
    
    return results


def get_rate_card_files_from_input():
    """
    Get all Excel files from the input folder that could be rate cards.
    
    Returns:
        list: List of Excel file names in the input folder
    """
    input_folder = "input"
    if not os.path.exists(input_folder):
        print(f"Warning: Input folder '{input_folder}' does not exist.")
        return []
    
    excel_files = [f for f in os.listdir(input_folder) 
                   if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    return excel_files


if __name__ == "__main__":
    # Example 1: Process multiple specific rate cards
    rate_card_files = ["ra_densir.xlsx"]
    results = process_multiple_rate_cards(rate_card_files)
    
    # Example 2: Process a single rate card (backward compatible)
    # output_file = save_rate_card_output("rate.xlsx")
    
    #Example 3: Auto-discover and process all Excel files in input folder
    #all_rate_cards = get_rate_card_files_from_input()
    # if all_rate_cards:
    #     results = process_multiple_rate_cards(all_rate_cards)
    
    # Default behavior: Process single file for testing
    #print("Processing single rate card (rate.xlsx)...")
    #output_file = save_rate_card_output("rate.xlsx")
    
    # Also print details to console
    #rate_card_dataframe, rate_card_column_names, rate_card_conditions, agreement_number = process_rate_card("rate.xlsx")
    #print("\nAgreement Number:", agreement_number if agreement_number else "Not found")
    #print("\nDataFrame shape:", rate_card_dataframe.shape)
    #print("\nColumn names:")
    #print(rate_card_column_names)
    #print("\nConditions (cleaned):")
    #for col, condition in rate_card_conditions.items():
    #    cleaned = clean_condition_text(condition)
    #    print(f"  {col}: {cleaned[:100]}..." if len(cleaned) > 100 else f"  {col}: {cleaned}")
    
    # Print Business Rules
    #print("\n" + "="*60)
    #print("BUSINESS RULES")
    #print("="*60)
    #business_rules = process_business_rules("rate.xlsx")
    #business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    #print(f"\nParsed {len(business_rules_conditions)} business rules:")
    #for rule_name, condition in business_rules_conditions.items():
    #    formatted = format_business_rule_condition(rule_name, condition)
    #    print(f"  {rule_name}: {formatte   d}")
    
    # Find and print which columns contain business rules
    #print("\n" + "="*60)
    #print("BUSINESS RULE COLUMNS IN RATE CARD")
    #print("="*60)
    #business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    #print(f"\nUnique columns containing business rule values:")
    #for col in sorted(business_rule_columns['unique_columns']):
    #    rules_count = len(business_rule_columns['column_to_rules'].get(col, []))
    #    print(f"  - {col}: {rules_count} rules")
